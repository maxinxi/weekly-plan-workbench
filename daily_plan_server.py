"""Local daily workbench. Workbook layout and summaries delegate to the supplied 5.6 code.

Never invoke legacy.main(): it scans the current directory and kills Excel processes.
Only disposable work copies are passed to its individual functions.
"""
from __future__ import annotations
import argparse
import base64
from contextlib import contextmanager
from collections import Counter, defaultdict
from datetime import datetime, date
import hashlib
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import importlib.util
import io
import json
import os
from pathlib import Path
import re
import secrets
import shutil
import sys
import tempfile
import threading
import time
import traceback
from urllib.parse import urlparse, parse_qs, quote
import webbrowser
import xml.etree.ElementTree as ET
from zipfile import ZipFile, ZIP_DEFLATED

ROOT = Path(__file__).resolve().parent
NS = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
LEGACY = None
ENGINE_LOCK = threading.RLock()


def legacy():
    global LEGACY
    if LEGACY is None:
        if os.environ.get('DAILY_LIBRARY_DIR'):
            sys.path.insert(0,os.environ['DAILY_LIBRARY_DIR'])
        spec = importlib.util.spec_from_file_location('daily_original_56', ROOT / 'daily_plan_legacy.py')
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        mod.ENABLE_BACKUP = False  # Our immutable input backup precedes all work-copy processing.
        mod.kill_excel_processes = lambda: (_ for _ in ()).throw(RuntimeError('禁止结束已有 Excel 进程'))
        mod.ask_yes_no_dialog = lambda *a, **kw: False
        mod.ask_yes_no = lambda *a, **kw: False
        mod.ask_text_dialog = lambda *a, **kw: ''
        def isolated_excel():
            # Do not fall back to an existing Excel instance owned by the user.
            app = mod.win32.DispatchEx('Excel.Application')
            # Select the A3-capable device for THIS instance; never change the Windows default.
            try:
                import win32api
                port = win32api.GetProfileVal('devices','Microsoft Print to PDF','').split(',')[-1]
                if port:
                    if 'Microsoft Print to PDF' not in str(app.ActivePrinter):
                        separator = ' 在 ' if ' 在 ' in str(app.ActivePrinter) else ' on '
                        app.ActivePrinter = 'Microsoft Print to PDF' + separator + port
            except Exception as e:
                mod.log('打印设备提示：使用 Excel 当前设备，A3 设置若失败将停止输出：' + str(e))
            return app
        mod.create_excel_application = isolated_excel
        # HTML owns all date decisions. No second dialog can silently change a title.
        mod.audit_plan_date_and_prompt_title_correction = lambda ws, path, title: title
        LEGACY = mod
    return LEGACY


def safe_name(name):
    if not isinstance(name, str) or not name or Path(name).name != name or re.search(r'[<>:"/\\|?*\x00-\x1f]', name):
        raise ValueError('文件名无效')
    if name.startswith('~$') or not name.lower().endswith('.xlsx'):
        raise ValueError('请导入 .xlsx 正式表（排除临时文件）')
    return name


def text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M')
    return str(value).replace('\r\n', '\n').replace('\r', '\n').strip()


def canonical(value):
    return re.sub(r'\s+', '', text(value)).replace('：', ':').replace('（', '(').replace('）', ')').lower()


def col_num(s):
    n = 0
    for ch in s.upper():
        n = n * 26 + ord(ch) - 64
    return n


def in_sqref(address, sqref):
    from openpyxl.utils.cell import range_boundaries, coordinate_from_string
    col, row = coordinate_from_string(address)
    for area in sqref.split():
        a, b, c, d = range_boundaries(area)
        if a <= col_num(col) <= c and b <= row <= d:
            return True
    return False


def validation_rules(data):
    """Read both standard and x14 validations without rewriting the input package."""
    with ZipFile(io.BytesIO(data)) as z:
        wb = ET.fromstring(z.read('xl/workbook.xml'))
        rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rid = wb.find('s:sheets', NS)[0].get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        target = next(r.get('Target') for r in rels if r.get('Id') == rid)
        path = target.lstrip('/') if target.startswith('/') else 'xl/' + target
        xml = ET.fromstring(z.read(path))
        rules = []
        for node in xml.iter():
            if node.tag.split('}')[-1] != 'dataValidation' or node.get('type') != 'list':
                continue
            ref = node.get('sqref') or next((text(n.text) for n in node.iter() if n.tag.split('}')[-1] == 'sqref'), '')
            fnode = next((n for n in node if n.tag.split('}')[-1] == 'formula1'), None)
            formula = ''.join(fnode.itertext()) if fnode is not None else ''
            if ref and formula:
                rules.append({'sqref': ref, 'formula': formula})
        return rules


def resolve_list(formula, wb, cached, row, depth=0):
    """Resolve direct ranges, defined names, inline lists and literal INDIRECT.

    Do not evaluate arbitrary Excel expressions or substitute a guessed personnel list.
    Excel itself remains the final formula authority at generation time.
    """
    if depth > 8:
        raise ValueError('下拉公式循环引用')
    f = formula.strip().lstrip('=')
    if f.startswith('"') and f.endswith('"'):
        return [x.strip() for x in f[1:-1].split(',') if x.strip()]
    m = re.fullmatch(r'INDIRECT\("([^"]+)"\)', f, re.I)
    if m:
        return resolve_list(m[1], wb, cached, row, depth + 1)
    name = wb.defined_names.get(f) or wb.worksheets[0].defined_names.get(f)
    if name:
        return resolve_list(name.attr_text, wb, cached, row, depth + 1)
    m = re.fullmatch(r"(?:'((?:[^']|'')+)'|([^!]+))!\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?", f, re.I)
    if not m:
        raise ValueError('此关联公式需要 Excel 计算；请使用“重新读取下拉公式”或手工填写')
    sheet = (m[1] or m[2]).replace("''", "'")
    if sheet not in wb.sheetnames:
        raise ValueError('下拉关联工作表不存在：' + sheet)
    a, b, c, d = col_num(m[3]), int(m[4]), col_num(m[5] or m[3]), int(m[6] or m[4])
    if (c-a+1)*(d-b+1) > 10000 or a > c or b > d:
        raise ValueError('下拉范围异常')
    result = []
    for rr in range(b, d+1):
        for cc in range(a, c+1):
            cell = wb[sheet].cell(rr, cc)
            v = cached[sheet].cell(rr, cc).value if cell.data_type == 'f' else cell.value
            if cell.data_type == 'f' and v is None:
                raise ValueError('名单单元格公式没有缓存值，需要 Excel 重新计算')
            v = text(v)
            if v and v not in result:
                result.append(v)
    return result


def read_document(name, data, kind):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = load_workbook(io.BytesIO(data), data_only=False)
        cached = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.worksheets[0]
    # Header content, not formatted UsedRange, determines the data columns.
    headers = {canonical(c.value): c.column for c in ws[2] if c.value is not None}
    def find(*parts):
        return next((col for h, col in headers.items() if any(p in h for p in parts)), None)
    fields = dict(work=find('工作内容'), start=find('计划开始', '开始时间'), end=find('计划竣工', '结束时间', '竣工时间'),
                  owner=find('负责人'), risk=find('作业风险'), grid=find('电网风险'))
    if kind == 'site':
        fields.update(same=find('同进同出人员'), profession=find('管理专业'), unit=find('检修(施工)单位'),
                      attendance=find('到岗到位人员'))
        # Never match the similarly named 人员层级 column.
        fields['same'] = next((c for h, c in headers.items() if h == '同进同出人员'), None)
    missing = [k for k, v in fields.items() if not v]
    if missing:
        labels = {'work':'工作内容','start':'计划开始时间','end':'计划竣工时间','owner':'负责人', 'same':'同进同出人员','risk':'作业风险等级','grid':'电网风险等级','profession':'管理专业','unit':'检修（施工）单位','attendance':'到岗到位人员'}
        raise ValueError(name + ' 缺少必需列：' + '、'.join(labels[x] for x in missing))
    rules = validation_rules(data) if kind == 'site' else []
    example_rows = [r for r in range(3, ws.max_row+1) if '例' in text(ws.cell(r,1).value)]
    example_rule = next((rule for r in example_rows for rule in rules if in_sqref(f'{get_column_letter(fields["same"])}{r}', rule['sqref'])), None) if kind == 'site' else None
    records = []
    for r in range(3, ws.max_row+1):
        # Merged continuation cells have no work value; they never count as another job.
        if r in example_rows or not text(ws.cell(r,fields['work']).value):
            continue
        record = {'id': f'{kind}:{r}', 'row':r, 'seq':text(ws.cell(r,1).value)}
        record.update({key:text(cached.worksheets[0].cell(r,col).value) for key,col in fields.items()})
        if kind == 'site':
            addr = f'{get_column_letter(fields["same"])}{r}'
            rule = next((x for x in rules if in_sqref(addr,x['sqref'])), example_rule)
            record.update(options=[], formula=rule['formula'] if rule else '', formulaError='', inherited=bool(rule and not in_sqref(addr,rule['sqref'])))
            if rule:
                try:
                    # Relative references track the first cell of each validation range,
                    # both within an existing multirow rule and when copying the example.
                    from openpyxl.formula.translate import Translator
                    origin_address = rule['sqref'].split()[0].split(':')[0]
                    record['formula'] = Translator('=' + rule['formula'].lstrip('='),origin=origin_address).translate_formula(addr).lstrip('=')
                    record['options'] = resolve_list(record['formula'],wb,cached,r)
                    if not record['options']:
                        record['formulaError'] = '关联范围没有可选人员'
                except Exception as e:
                    record['formulaError'] = str(e)
            else:
                record['formulaError'] = '本行及示例行均未找到下拉关联，请手工填写或修好源表后重新导入'
        records.append(record)
    result = {'name': name, 'kind':kind, 'title':text(ws.cell(1,1).value), 'fields': fields,
              'records':records, 'exampleRows':example_rows, 'hiddenSheets':[s.title for s in wb if s.sheet_state != 'visible']}
    wb.close(); cached.close()
    if not records:
        raise ValueError(name + ' 未找到有效作业记录')
    return result


def parse_time(value):
    s = text(value).replace('：', ':')
    m = re.search(r'(\d{4})[年/.-](\d{1,2})[月/.-](\d{1,2})日?\s*[T\s]*(\d{1,2}):(\d{1,2})(?::\d{1,2})?', s)
    if not m:
        return None, '时间缺日期或日期/时间格式无法解析'
    try:
        return datetime(*map(int,m.groups())), ''
    except ValueError:
        return None, '日期非法或时间超出有效范围'


def audit(documents):
    issues = []
    def issue(code,message,kind='',row=None,field='',value='',suggestion=None):
        item = dict(code=code,message=message,kind=kind,row=row,field=field,value=value,suggestion=suggestion)
        item['id'] = hashlib.sha256(json.dumps(item,ensure_ascii=False,sort_keys=True).encode()).hexdigest()[:18]
        issues.append(item)
    for kind,doc in documents.items():
        title_md = re.search(r'(\d{1,2})月(\d{1,2})日',doc['title'])
        file_md = re.search(r'(\d{1,2})月(\d{1,2})日',doc['name'])
        if title_md and file_md and title_md.group() != file_md.group():
            proposed = doc['title'][:title_md.start()] + file_md.group() + doc['title'][title_md.end():]
            issue('TITLE_DATE','标题日期与文件名不一致',kind,1,'title',doc['title'],proposed)
        try:
            original_check = legacy().evaluate_plan_date_consistency(doc['name'],doc['title'])
            if original_check['title_status'] not in legacy().PLAN_DATE_ACCEPTED_STATUSES:
                proposed = legacy().replace_title_plan_date(doc['title'],original_check['suggested_date']) if original_check['needs_title_change'] else None
                issue('PLAN_DATE',original_check['reason'],kind,1,'title',doc['title'],proposed)
        except (ValueError,TypeError):
            issue('TITLE_INVALID','标题或文件名中的日期非法',kind,1,'title',doc['title'])
        for record in doc['records']:
            dates = []
            for field in ('start','end'):
                dt,err = parse_time(record[field]); dates.append(dt)
                if err:
                    issue('TIME_INVALID',err,kind,record['row'],field,record[field])
                elif title_md and (dt.month,dt.day) != tuple(map(int,title_md.groups())):
                    issue('ROW_DATE','作业日期与标题日期不一致，请核实跨日计划',kind,record['row'],field,record[field])
            if all(dates) and dates[1] <= dates[0]:
                issue('TIME_ORDER','计划竣工时间不晚于开始时间',kind,record['row'],'end',record['end'])
            if record.get('formulaError'):
                issue('DROPDOWN',record['formulaError'],kind,record['row'],'same',record.get('same',''))
    if 'site' in documents and 'risk' in documents:
        site,risk = documents['site']['records'],documents['risk']['records']
        if len(site) != len(risk):
            issue('COUNT',f'现场作业 {len(site)} 项，风险管控 {len(risk)} 项；请核对缺项/重复项，程序不代替新增或删除作业')
        groups = [defaultdict(list),defaultdict(list)]
        for group,records in zip(groups,(site,risk)):
            for r in records:
                group[canonical(r['work'])].append(r)
        for key in set(groups[0]) | set(groups[1]):
            left,right = groups[0][key],groups[1][key]
            if len(left) == len(right) == 1:
                a,b = left[0],right[0]
                for field in ('start','end'):
                    av,_ = parse_time(a[field]); bv,_ = parse_time(b[field])
                    if (av and bv and av != bv) or (not (av and bv) and canonical(a[field]) != canonical(b[field])):
                        issue('CROSS_TIME',f'同一作业两表时间不一致（现场第 {a["row"]} 行 / 风险第 {b["row"]} 行）；建议值来自现场表', 'risk',b['row'],field,b[field],a[field])
            elif not left or not right:
                for rec in left or right:
                    issue('UNMATCHED','另一张表未找到相同工作内容，请人工核对；不会根据相近文字强行配对','site' if left else 'risk',rec['row'],'work',rec['work'])
            else:
                # Repeated descriptions are only paired when owner AND full start/end agree.
                signature = lambda r:(canonical(r['owner']),r['start'],r['end'])
                if Counter(map(signature,left)) != Counter(map(signature,right)):
                    issue('AMBIGUOUS','相同工作内容出现多次且时间或负责人不一致，无法唯一配对，请人工核对','site',left[0]['row'],'work',left[0]['work'])
    return issues


def summaries(doc, mode_b=False):
    mod = legacy()
    records = []
    missing = 0
    for i,r in enumerate(doc['records'],1):
        start,_ = parse_time(r['start']); end,_ = parse_time(r['end'])
        same = mod.remove_phone_and_noise(r['same']) if r['same'] else ''
        if not same:
            missing += 1
            same = f'需要领导{missing}'
        records.append(dict(order=i,source_row=r['row'],seq=str(i),profession=mod.classify_profession(r['profession'],r['owner']),raw_profession=r['profession'],work=r['work'],unit=r['unit'],leader=mod.clean_person_name(r['owner']),time_part=f'{start:%H:%M} - {end:%H:%M}' if start and end else '',risk=r['risk'],grid_risk='' if r['grid'] == '无' else r['grid'],same_inout=same,need_leader_no=missing if not r['same'] else None))
    mod.prepare_summary_source_entries(records)
    result = {'a':'\n'.join(mod.build_mode_a_summary_lines(doc['title'],records)), 'b':''}
    if mode_b:
        if missing:
            raise ValueError('同进同出人员尚未全部填写')
        result['b'] = '\n'.join(mod.build_mode_b_summary_lines(doc['title'],records,Counter(r['risk'] for r in records)))
    return result


def backup_once(path, data):
    path.parent.mkdir(parents=True,exist_ok=True)
    try:
        with path.open('xb') as f:
            f.write(data)
    except FileExistsError:
        pass


def patch_package(data, doc, original):
    """Edit only explicit values + missing list rules. Preserve x14/hidden sheets/styles.

    No openpyxl save round-trip: it discards Excel extended data validations.
    """
    from xml.sax.saxutils import escape
    from openpyxl.utils import get_column_letter
    updates = {}
    if doc['title'] != original['title']:
        updates['A1'] = doc['title']
    original_rows = {r['row']:r for r in original['records']}
    for r in doc['records']:
        for field,col in doc['fields'].items():
            if r[field] != original_rows[r['row']][field]:
                updates[f'{get_column_letter(col)}{r["row"]}'] = r[field]
    # Determine the first sheet target without assuming sheet1.xml.
    result = io.BytesIO()
    with ZipFile(io.BytesIO(data)) as zin, ZipFile(result,'w',ZIP_DEFLATED) as zout:
        w = ET.fromstring(zin.read('xl/workbook.xml'))
        rid = w.find('s:sheets',NS)[0].get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        target = next(r.get('Target') for r in ET.fromstring(zin.read('xl/_rels/workbook.xml.rels')) if r.get('Id') == rid)
        sheetpath = target.lstrip('/') if target.startswith('/') else 'xl/' + target
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename == sheetpath:
                xml = content.decode('utf-8')
                for address,value in updates.items():
                    pattern = rf'<c\b(?=[^>]*\br="{address}")[^>]*(?:/>|>[\s\S]*?</c>)'
                    match = re.search(pattern,xml)
                    attrs = re.match(r'<c\b([^>]*?)(?:/?>)',match[0])[1] if match else f' r="{address}"'
                    attrs = re.sub(r'\s+t="[^"]*"','',attrs).rstrip('/')
                    cell = f'<c{attrs} t="inlineStr"><is><t xml:space="preserve">{escape(value)}</t></is></c>'
                    if match:
                        xml = xml[:match.start()] + cell + xml[match.end():]
                    else:
                        row = re.search(r'\d+',address)[0]
                        rowpat = rf'(<row\b[^>]*\br="{row}"[^>]*>)([\s\S]*?)(</row>)'
                        if not re.search(rowpat,xml):
                            raise ValueError('找不到需修改的原表行 ' + row)
                        def append_cell(m):
                            cells = re.findall(r'<c\b[^>]*(?:/>|>[\s\S]*?</c>)',m[2]) + [cell]
                            cells.sort(key=lambda x: col_num(re.search(r'\br="([A-Z]+)',x)[1]))
                            return m[1] + ''.join(cells) + m[3]
                        xml = re.sub(rowpat,append_cell,xml,count=1)
                if doc['kind'] == 'site':
                    extra = []
                    for r in doc['records']:
                        if r.get('inherited') and r['formula']:
                            addr = f'{get_column_letter(doc["fields"]["same"])}{r["row"]}'
                            extra.append(f'<dataValidation type="list" allowBlank="1" showErrorMessage="0" sqref="{addr}"><formula1>{escape(r["formula"])}</formula1></dataValidation>')
                    if extra:
                        m = re.search(r'<dataValidations\b[^>]*>([\s\S]*?)</dataValidations>',xml)
                        inside = (m[1] if m else '') + ''.join(extra)
                        tag = f'<dataValidations count="{len(re.findall(r"<dataValidation\b",inside))}">{inside}</dataValidations>'
                        if m:
                            xml = xml[:m.start()] + tag + xml[m.end():]
                        else:
                            pos = re.search(r'<(?:hyperlinks|printOptions|pageMargins|pageSetup|headerFooter|rowBreaks|colBreaks|drawing|extLst)\b|</worksheet>',xml).start()
                            xml = xml[:pos] + tag + xml[pos:]
                content = xml.encode('utf-8')
            zout.writestr(item,content)
    return result.getvalue()


def serial_only_rows(data, kind):
    """Find genuinely empty numbered rows, protecting real merged continuations."""
    from openpyxl import load_workbook
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        book = load_workbook(io.BytesIO(data),data_only=False)
    sheet = book.worksheets[0]
    last_col = 20 if kind == 'daogang' else 19 if kind == 'site' else 12
    sidebars = {12,13} if kind == 'daogang' else set()
    result = []
    for row in range(3,sheet.max_row+1):
        if not re.fullmatch(r'\d+(?:\.0)?',text(sheet.cell(row,1).value)):
            continue
        empty = True
        for col in range(2,last_col+1):
            if col in sidebars:
                continue
            value = sheet.cell(row,col).value
            if text(value):
                empty = False; break
            # A continuation can display content whose anchor sits in an earlier row.
            for merge in sheet.merged_cells.ranges:
                if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
                    if text(sheet.cell(merge.min_row,merge.min_col).value):
                        empty = False
                    break
            if not empty:
                break
        if empty:
            result.append(row)
    book.close()
    return result


def excel_export(paths, options):
    """Use native Excel PDF rendering; adjust only existing border segments on request."""
    mod = legacy()
    excel = mod.create_excel_application()
    mod.configure_excel_silent(excel)
    cleaned = []
    try:
        for path in paths:
            kind = 'daogang' if '到岗到位' in path.name else 'risk' if '风险管控' in path.name else 'site'
            original_bytes = path.read_bytes()
            empty_rows = serial_only_rows(original_bytes,kind)
            # COM can expose a merged anchor through continuation cells. Capture actual
            # XML work anchors before editing so each job receives exactly one number.
            work_rows = [r['row'] for r in read_document(path.name,original_bytes,kind)['records']] if empty_rows else []
            wb = excel.Workbooks.Open(str(path.resolve()),UpdateLinks=0,ReadOnly=False,IgnoreReadOnlyRecommended=True)
            try:
                ws = wb.Worksheets(1)
                if empty_rows:
                    last = mod.get_used_rows(ws) - len(empty_rows)
                    for row in reversed(empty_rows):
                        ws.Rows(row).Delete()
                    # Number only actual work anchors, never manufacture a new blank record.
                    for sequence,old_row in enumerate(work_rows,1):
                        row = old_row - sum(deleted < old_row for deleted in empty_rows)
                        cell = ws.Cells(row,1)
                        anchor = cell.MergeArea.Cells(1,1) if cell.MergeCells else cell
                        anchor.Value = sequence
                    print_col = 12 if kind == 'risk' else 19
                    ws.PageSetup.PrintArea = ws.Range(ws.Cells(1,1),ws.Cells(last,print_col)).Address
                    wb.Save()
                    cleaned.append({'file':path.name,'rows':empty_rows})
                    mod.log(f'已删除只有序号的空白行：{path.name}，原排版第 {empty_rows} 行；有效作业序号已重新核对，打印区域已收口')
                weight = options.get('border','original')
                color = options.get('borderColor','#000000')
                if kind != 'site' and (weight != 'original' or color != '#000000'):
                    if weight not in ('original','hairline','thin','medium') or not re.fullmatch(r'#[0-9a-fA-F]{6}',color):
                        raise ValueError('框线设置无效')
                    rgb = int(color[1:3],16) + (int(color[3:5],16)<<8) + (int(color[5:7],16)<<16)
                    last_col = 19 if '到岗到位' in path.name else 12
                    visited = set()
                    for row in range(2,mod.get_used_rows(ws)+1):
                        for col in range(1,last_col+1):
                            cell = ws.Cells(row,col)
                            area = cell.MergeArea if cell.MergeCells else cell
                            address = str(area.Address)
                            if address in visited:
                                continue
                            visited.add(address)
                            for edge in (7,8,9,10):
                                border = area.Borders(edge)
                                if border.LineStyle != -4142:
                                    if weight != 'original':
                                        border.Weight = {'hairline':1,'thin':2,'medium':-4138}[weight]
                                    border.Color = rgb
                    wb.Save()
                if kind != 'site':
                    wb.ExportAsFixedFormat(0,str(path.with_suffix('.pdf').resolve()),0,True,False)
                    if not path.with_suffix('.pdf').is_file():
                        raise RuntimeError('Excel 未生成 PDF：' + path.name)
            finally:
                wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return cleaned


@contextmanager
def a3_printer():
    """Excel derives printable A3 metrics from the printer driver at instance creation.

    Some localized Office versions reject ActivePrinter assignments. A short-lived
    default-device switch lets Excel use the A3 driver; always restore it afterwards.
    """
    import win32print
    previous = win32print.GetDefaultPrinter()
    changed = previous != 'Microsoft Print to PDF'
    if changed:
        win32print.SetDefaultPrinter('Microsoft Print to PDF')
    try:
        yield
    finally:
        if changed and win32print.GetDefaultPrinter() == 'Microsoft Print to PDF':
            win32print.SetDefaultPrinter(previous)


class Session:
    def __init__(self, base):
        self.base = Path(base)
        self.inputs = {}; self.original = {}; self.documents = {}; self.changes = []
        self.result = {}; self.status = {'busy':False,'phase':'等待导入','log':[]}
        self.revision = 0
        self.source_paths = {}

    def load(self, files):
        docs = {}; inputs = {}
        if len(files) != 2:
            raise ValueError('请同时导入现场作业计划和风险管控，两份 .xlsx 文件')
        for file in files:
            name = safe_name(file['name'])
            data = base64.b64decode(file['data'],validate=True)
            if len(data)>30*1024*1024:
                raise ValueError('单份文件超过 30 MB')
            kind = 'risk' if '风险管控' in name else 'site'
            if kind in docs or '到岗到位' in name:
                raise ValueError('请导入一份现场源表和一份风险管控，排除已生成的到岗到位文件')
            docs[kind] = read_document(name,data,kind); inputs[kind] = data
        if set(docs) != {'site','risk'}:
            raise ValueError('未识别到两类输入文件')
        for kind,data in inputs.items():
            digest = hashlib.sha256(data).hexdigest()
            backup_once(self.base/'源文件备份'/digest/docs[kind]['name'],data)
        self.inputs = inputs
        self.original = json.loads(json.dumps(docs,ensure_ascii=False))
        self.documents = docs; self.changes = []; self.result = {}; self.revision += 1
        self.source_paths = {}
        return self.view()

    def load_paths(self, paths):
        """Paths come only from the user-operated native file picker (or explicit CLI tests)."""
        files = []
        chosen = {}
        for raw in paths:
            p = Path(raw).resolve(strict=True)
            if any(part in ('（处理后）','源文件备份') or part.startswith('work-') for part in p.parts):
                raise ValueError('请选择原始目录中的源表，排除（处理后）、源文件备份和 work 副本')
            safe_name(p.name)
            if p.stat().st_size > 30*1024*1024:
                raise ValueError('单份文件超过 30 MB')
            data = p.read_bytes()
            kind = 'risk' if '风险管控' in p.name else 'site'
            chosen[kind] = p
            files.append({'name':p.name,'data':base64.b64encode(data).decode()})
        self.load(files)
        for kind,p in chosen.items():
            backup_once(p.parent/'源文件备份'/p.name,self.inputs[kind])
        self.source_paths = chosen
        return self.view()

    def view(self):
        return {'documents':self.documents,'issues':audit(self.documents),'revision':self.revision,
                'ready':bool(self.documents) and all(r['same'].strip() for r in self.documents['site']['records']),
                'changes':self.changes,
                'outputLocation':str(self.source_paths['site'].parent/'（处理后）') if self.source_paths else '本机输出目录（兼容导入，可下载完整压缩包）'}

    def update(self, edits):
        allowed = {'same','start','end','work','title'}
        for edit in edits:
            kind,field,row = edit['kind'],edit['field'],edit.get('row')
            if kind not in self.documents or field not in allowed or (field == 'same' and kind != 'site'):
                raise ValueError('此字段不允许从网页修改')
            value = str(edit['value'])
            if len(value)>32767:
                raise ValueError('单元格文字超过 Excel 上限')
            target = self.documents[kind] if field == 'title' else next(r for r in self.documents[kind]['records'] if r['row'] == row)
            before = target[field]
            if before != value:
                target[field] = value
                self.changes.append(dict(kind=kind,row=row,field=field,before=before,after=value))
                self.revision += 1
        self.result = {}
        return self.view()

    def refresh_dropdowns(self):
        import pythoncom
        pythoncom.CoInitialize()
        app = None
        try:
            app = legacy().create_excel_application()
            legacy().configure_excel_silent(app)
            with tempfile.TemporaryDirectory(prefix='work-list-',dir=self.base) as temp:
                path = Path(temp)/self.documents['site']['name']
                path.write_bytes(self.inputs['site'])
                book = app.Workbooks.Open(str(path),UpdateLinks=0,ReadOnly=True)
                try:
                    sheet = book.Worksheets(1)
                    sheet.Activate()
                    app.Calculate()
                    for r in self.documents['site']['records']:
                        if not r['formula']:
                            continue
                        try:
                            sheet.Cells(r['row'],self.documents['site']['fields']['same']).Activate()
                            value = app.Evaluate('=' + r['formula'].lstrip('='))
                            if hasattr(value,'Value'):
                                value = value.Value
                            if not isinstance(value,(tuple,list)):
                                value = [value]
                            flat = []
                            for item in value:
                                for entry in item if isinstance(item,(tuple,list)) else [item]:
                                    if isinstance(entry,(int,float)) and entry < -100000:
                                        raise ValueError('Excel 无法解析关联公式')
                                    v = text(entry)
                                    if v and v not in flat:
                                        flat.append(v)
                            if not flat:
                                raise ValueError('关联范围没有人员')
                            r['options'],r['formulaError'] = flat,''
                        except Exception as e:
                            r['formulaError'] = 'Excel 计算关联公式失败：' + str(e)
                finally:
                    book.Close(SaveChanges=False)
            return self.view()
        finally:
            if app:
                app.Quit()
            pythoncom.CoUninitialize()

    def generate(self, options):
        import pythoncom
        with ENGINE_LOCK:
            pythoncom.CoInitialize()
            mod = legacy()
            printer_context = a3_printer()
            printer_entered = False
            oldlog,oldred = mod.log,mod.log_red
            original_config = {k:getattr(mod,k) for k in ('RISK_AUTOFIT_PLUS_HEIGHT','RISK_COMPACT_AUTOFIT_PLUS_HEIGHT','DEFAULT_AUTOFIT_PLUS_HEIGHT')}
            try:
                self.status = {'busy':True,'phase':'正在建立工作副本','log':[],'error':''}
                def log(message=''):
                    message = text(message)
                    self.status['phase'] = message[-240:]
                    self.status['log'].append(message)
                    mod.REPORT_LINES.append(message)
                mod.log = log; mod.log_red = log; mod.REPORT_LINES.clear()
                if not self.documents or not all(r['same'].strip() for r in self.documents['site']['records']):
                    raise ValueError('有效作业的同进同出人员必须全部填写')
                current_issues = audit(self.documents)
                if any(i['id'] not in options.get('acknowledged',[]) for i in current_issues):
                    raise ValueError('仍有未确认的告警，请先修正或选择保留告警继续')
                printer_context.__enter__()
                printer_entered = True
                plus = options.get('extraHeight')
                if plus is not None:
                    plus = float(plus)
                    if not 0.5 <= plus <= 20:
                        raise ValueError('额外行高应在 0.5—20 磅之间')
                    for key in original_config:
                        setattr(mod,key,plus)
                output = self.base/'results'/secrets.token_hex(6)/'（处理后）'
                output.mkdir(parents=True)
                self.result = {}
                # TemporaryDirectory cleans work on success AND failure. Inputs never enter legacy code.
                with tempfile.TemporaryDirectory(prefix='work-',dir=self.base) as temp:
                    work = Path(temp)
                    paths = {}
                    for kind,doc in self.documents.items():
                        path = work/doc['name']
                        path.write_bytes(patch_package(self.inputs[kind],doc,self.original[kind]))
                        paths[kind] = path
                        result = mod.preprocess_excel(str(path),str(work),mode=kind)
                        if result in {'文件被占用','只读无法修改','处理失败','文件不存在','缓存错误'}:
                            raise RuntimeError(doc['name'] + '：' + result)
                    target = Path(mod.build_daogang_path(str(paths['site'])))
                    shutil.copy2(paths['site'],target)
                    if not mod.transform_daogang_columns(str(target)):
                        raise RuntimeError('到岗到位列转换失败')
                    side = options.get('sidebar',{})
                    operations = '\n'.join(f'{label}：{mod.normalize_personnel_multiline_text(side[key])}' for key,label in [('operations','运检专业'),('marketing','营销专业')] if side.get(key,'').strip())
                    sidebar = {'operations':operations,'leader':mod.format_company_leader_text(side.get('leader',''))}
                    result = mod.preprocess_excel(str(target),str(work),mode='daogang',daogang_sidebar=sidebar)
                    if result in {'文件被占用','只读无法修改','处理失败','文件不存在','缓存错误'}:
                        raise RuntimeError('到岗到位排版失败：' + result)
                    log('按原 Excel 打印设置导出两个 PDF…')
                    cleaned_rows = excel_export([paths['site'],target,paths['risk']],options)
                    final_docs = {kind:read_document(path.name,path.read_bytes(),kind) for kind,path in paths.items()}
                    if any(len(final_docs[k]['records']) != len(self.documents[k]['records']) for k in paths):
                        raise RuntimeError('处理前后有效作业数量发生变化，已停止交付，请查看处理报告')
                    from openpyxl import load_workbook
                    final_warnings = []
                    import warnings
                    for kind,path in paths.items():
                        with warnings.catch_warnings():
                            warnings.simplefilter('ignore')
                            checked = load_workbook(path,data_only=True)
                        sheet = checked.worksheets[0]
                        work_col = final_docs[kind]['fields']['work']
                        for rr in range(3,sheet.max_row+1):
                            if text(sheet.cell(rr,1).value) and not text(sheet.cell(rr,work_col).value):
                                final_warnings.append(f'{path.name} 第 {rr} 行有序号但缺工作内容，仍含其他数据或合并内容，未按空白行删除；请核对缺填内容。')
                        checked.close()
                    # The final saved site workbook, not a stale Mode A preview, drives both summaries.
                    summary,counter = mod.process_excel(str(paths['site']))
                    mod.prepare_summary_source_entries(summary['records'])
                    texts = {'a':'\n'.join(mod.build_mode_a_summary_lines(summary['title'],summary['records'])),
                             'b':'\n'.join(mod.build_mode_b_summary_lines(summary['title'],summary['records'],counter))}
                    for path in [paths['site'],target,paths['risk'],target.with_suffix('.pdf'),paths['risk'].with_suffix('.pdf')]:
                        shutil.copy2(path,output/path.name)
                    (output/'汇总-模式A.txt').write_text(texts['a'],encoding='utf-8-sig')
                    (output/'汇总-模式B.txt').write_text(texts['b'],encoding='utf-8-sig')
                    (output/'核对与修改报告.json').write_text(json.dumps({'issues':current_issues,'finalWarnings':final_warnings,'removedSerialOnlyRows':cleaned_rows,'acknowledged':options.get('acknowledged',[]),'changes':self.changes,'counts':{k:len(v['records']) for k,v in self.documents.items()},'options':options},ensure_ascii=False,indent=2),encoding='utf-8')
                    (output/'原版处理报告.txt').write_text('\n'.join(self.status['log']),encoding='utf-8-sig')
                zip_path = output.parent/'（处理后）.zip'
                with ZipFile(zip_path,'w',ZIP_DEFLATED) as z:
                    for p in output.iterdir():
                        z.write(p,'（处理后）/'+p.name)
                output_names = [p.name for p in output.iterdir()]
                if self.source_paths:
                    destination = self.source_paths['site'].parent/'（处理后）'
                    publish_output(output,destination)
                    output = destination
                self.result = {'folder':str(output),'zip':str(zip_path),'files':output_names, 'summaries':texts,'revision':self.revision,'warnings':final_warnings}
                self.status['phase'] = '三份 Excel、两个 PDF 和模式 A/B 汇总已生成；work 副本已清理'
            except Exception as e:
                self.status['error'] = str(e)
                self.status['phase'] = '处理失败，原表未修改，work 副本已清理'
                self.status['log'].append(traceback.format_exc())
            finally:
                mod.log,mod.log_red = oldlog,oldred
                for k,v in original_config.items():
                    setattr(mod,k,v)
                try:
                    if printer_entered:
                        printer_context.__exit__(None,None,None)
                finally:
                    self.status['busy'] = False
                    pythoncom.CoUninitialize()


def publish_output(source, destination):
    """Stage all replacements before publishing; restore previous files if one is locked."""
    destination = Path(destination)
    destination.mkdir(parents=True,exist_ok=True)
    if destination.resolve() == Path(source).resolve():
        return
    with tempfile.TemporaryDirectory(prefix='work-publish-',dir=destination.parent) as tmp:
        temp = Path(tmp); staged=temp/'new';old=temp/'old';staged.mkdir();old.mkdir()
        files = [p for p in Path(source).iterdir() if p.is_file()]
        for p in files:
            shutil.copy2(p,staged/p.name)
            target = destination/p.name
            if target.exists():
                shutil.copy2(target,old/p.name)
        installed = []
        try:
            for p in files:
                os.replace(staged/p.name,destination/p.name)
                installed.append(p.name)
        except Exception:
            for name in reversed(installed):
                if (old/name).exists():
                    os.replace(old/name,destination/name)
                else:
                    (destination/name).unlink()
            raise


def serve(port=0, open_browser=True, base=None):
    token = secrets.token_urlsafe(32)
    base = Path(base) if base else Path(os.environ.get('LOCALAPPDATA',tempfile.gettempdir()))/'DailyPlanWorkbench'
    base.mkdir(parents=True,exist_ok=True)
    session = Session(base)
    class Handler(BaseHTTPRequestHandler):
        def log_message(self,*args):
            pass

        def send(self, data, status=200, mime='application/json; charset=utf-8', extra=None):
            if isinstance(data,(dict,list)):
                data = json.dumps(data,ensure_ascii=False).encode('utf-8')
            elif isinstance(data,str):
                data = data.encode('utf-8')
            self.send_response(status)
            self.send_header('Content-Type',mime)
            self.send_header('Content-Length',str(len(data)))
            self.send_header('Cache-Control','no-store')
            self.send_header('X-Content-Type-Options','nosniff')
            self.send_header('Referrer-Policy','no-referrer')
            for k,v in (extra or {}).items():
                self.send_header(k,v)
            self.end_headers(); self.wfile.write(data)

        def authorized(self):
            host = f'127.0.0.1:{server.server_port}'
            if self.headers.get('Host') != host:
                return False
            cookies = self.headers.get('Cookie','').split(';')
            return any(secrets.compare_digest(x.strip(),'daily_token='+token) for x in cookies)

        def do_GET(self):
            parsed = urlparse(self.path)
            query = parse_qs(parsed.query)
            if parsed.path == '/' and secrets.compare_digest(query.get('token',[''])[0],token):
                self.send('',302,extra={'Location':'/','Set-Cookie':f'daily_token={token}; HttpOnly; SameSite=Strict; Path=/'})
                return
            if not self.authorized():
                self.send({'error':'请从本机启动器打开工作台'},403); return
            if parsed.path == '/':
                html = (ROOT/'日计划工作台.html').read_text(encoding='utf-8')
                self.send(html.replace('/*LOCAL_CONNECTION*/',f'window.LOCAL_TOKEN={json.dumps(token)};'),mime='text/html; charset=utf-8')
            elif parsed.path == '/api/status':
                self.send({'status':session.status,'result':{k:v for k,v in session.result.items() if k not in ('folder','zip')}})
            elif parsed.path == '/api/state':
                self.send(session.view() if session.documents else {'empty':True})
            elif parsed.path == '/api/file':
                name = query.get('name',[''])[0]
                if name == '（处理后）.zip' and session.result:
                    path = Path(session.result['zip'])
                elif session.result and name in session.result['files']:
                    path = Path(session.result['folder'])/name
                else:
                    self.send({'error':'文件不存在'},404); return
                mime = 'application/pdf' if path.suffix == '.pdf' else 'application/octet-stream'
                disposition = 'inline' if path.suffix == '.pdf' and query.get('preview') else 'attachment'
                self.send(path.read_bytes(),mime=mime,extra={'Content-Disposition':f"{disposition}; filename*=UTF-8''{quote(name)}"})
            else:
                self.send({'error':'不存在'},404)

        def do_POST(self):
            origin = f'http://127.0.0.1:{server.server_port}'
            if not self.authorized() or self.headers.get('Origin') != origin or not secrets.compare_digest(self.headers.get('X-Daily-Token',''),token):
                self.send({'error':'本机访问校验失败'},403); return
            try:
                size = int(self.headers.get('Content-Length','0'))
                if not 0 <= size <= 85*1024*1024:
                    raise ValueError('请求过大')
                data = json.loads(self.rfile.read(size) or b'{}')
                if session.status['busy']:
                    self.send({'error':'正在处理，请稍候'},409); return
                with ENGINE_LOCK:
                    if self.path == '/api/import':
                        result = session.load(data['files'])
                    elif self.path == '/api/select':
                        import tkinter as tk
                        from tkinter import filedialog
                        picker = tk.Tk(); picker.withdraw(); picker.attributes('-topmost',True)
                        try:
                            paths = filedialog.askopenfilenames(parent=picker,title='选择现场作业计划和风险管控（两份源表）',filetypes=[('Excel 工作簿','*.xlsx')])
                        finally:
                            picker.destroy()
                        result = session.load_paths(paths) if paths else {'cancelled':True}
                    elif self.path == '/api/update':
                        result = session.update(data['edits'])
                    elif self.path == '/api/summary':
                        result = summaries(session.documents['site'],False)
                    elif self.path == '/api/dropdowns':
                        result = session.refresh_dropdowns()
                    elif self.path == '/api/generate':
                        session.status['busy'] = True
                        threading.Thread(target=session.generate,args=(data,),daemon=True).start()
                        result = {'started':True}
                    elif self.path == '/api/folder':
                        if session.result:
                            os.startfile(session.result['folder'])
                        result = {'ok':True}
                    elif self.path == '/api/stop':
                        threading.Thread(target=server.shutdown,daemon=True).start()
                        result = {'ok':True}
                    else:
                        raise ValueError('未知请求')
                self.send(result)
            except Exception as e:
                self.send({'error':str(e)},400)
    server = ThreadingHTTPServer(('127.0.0.1',port),Handler)
    url = f'http://127.0.0.1:{server.server_port}/?token={token}'
    print('日计划工作台已启动：'+url,flush=True)
    if open_browser:
        webbrowser.open(url)
    try:
        server.serve_forever()
    finally:
        server.server_close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='日计划本机工作台：原版 Excel 排版、模式 A/B、PDF 导出')
    parser.add_argument('--port',type=int,default=0)
    parser.add_argument('--no-browser',action='store_true')
    parser.add_argument('--base',type=Path)
    args = parser.parse_args()
    try:
        import openpyxl, pandas, win32com.client
        serve(args.port,not args.no_browser,args.base)
    except Exception:
        message = traceback.format_exc()
        (ROOT/'启动错误.txt').write_text(message,encoding='utf-8')
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0,'启动失败。请查看同目录“启动错误.txt”。\n需要 Windows、Microsoft Excel 和 Python 的 openpyxl、pandas、pywin32。\n\n'+message[-1600:],'日计划工作台',16)
        except Exception:
            print(message)
