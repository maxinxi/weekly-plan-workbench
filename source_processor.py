"""表1处理：只读原件 → 一次备份 → work → 清洗 → 列宽 → 行高 → 插行。"""
from __future__ import annotations

import argparse
from contextlib import contextmanager
from copy import copy
from datetime import date, datetime, timedelta
import json
import math
from pathlib import Path
import re
import shutil
import tempfile
import unicodedata
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.pagebreak import Break, RowBreak, ColBreak
from openpyxl.utils import get_column_letter, range_boundaries

MEASURES_HEIGHT_PLUS = 1.0
MAX_HEIGHT = 409.5
KEYS = {
    "content": ("作业内容", "工作内容", "作业项目", "作业任务"),
    "time": ("作业时间", "工作时间", "计划时间", "起止时间", "时间"),
    "measures": ("管控措施", "控制措施", "防控措施", "风险措施"),
    "person": ("同进同出", "负责人", "人员", "联系人", "姓名"),
    "date": ("日期",),
    "risk": ("作业风险等级", "风险等级", "风险级别"),
}
DATE = r"(\d{4})[年/.-](\d{1,2})[月/.-](\d{1,2})日?"
CLOCK = r"(\d{1,2}):(\d{2})"


def trim_inflated_copy(path, report):
    """Remove only far-right blank formatting from the work copy before Excel loads it.

    Some Excel exports contain 40,000 empty merges out to XFC. Keep all actual
    values/formulas and all merges intersecting the data columns, without
    reserializing XML namespaces or touching any other worksheet.
    """
    raw = Path(path).read_bytes()
    with ZipFile(BytesIO(raw)) as source:
        book = ET.fromstring(source.read('xl/workbook.xml'))
        sheet = book.find('{*}sheets/{*}sheet')
        rid = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        rels = ET.fromstring(source.read('xl/_rels/workbook.xml.rels'))
        target = next(r.get('Target') for r in rels if r.get('Id') == rid)
        name = target.lstrip('/') if target.startswith('/') else 'xl/' + target
        xml = source.read(name).decode('utf-8')
        cells = list(re.finditer(r'<c\b[^>]*?(?:/>|>.*?</c>)', xml, re.S))
        def coord(cell):
            return range_boundaries(re.search(r'\br="([A-Z]+\d+)"', cell)[1])[:2]
        populated = [coord(m[0])[0] for m in cells if re.search(r'<(?:v|is|f)(?:\s|>)', m[0])]
        if not populated:
            return
        bound = max(populated)
        ranges = re.findall(r'<mergeCell\b[^>]*\bref="([^"]+)"[^>]*/>', xml)
        for ref in ranges:
            l, _, r, _ = range_boundaries(ref)
            if l <= bound:
                bound = max(bound, r)
        actual_bound = max((coord(m[0])[0] for m in cells), default=bound)
        if actual_bound <= bound + 256:
            return
        xml = re.sub(r'<c\b[^>]*?(?:/>|>.*?</c>)', lambda m: m[0] if coord(m[0])[0] <= bound else '', xml, flags=re.S)
        kept = [ref for ref in ranges if range_boundaries(ref)[0] <= bound]
        xml = re.sub(r'<mergeCells\b[^>]*>.*?</mergeCells>', lambda _: '<mergeCells count="%d">%s</mergeCells>' % (len(kept), ''.join('<mergeCell ref="%s"/>' % ref for ref in kept)), xml, flags=re.S)
        # Keep the column attributes verbatim except the right endpoint.
        def trim_col(m):
            lo = int(re.search(r'\bmin="(\d+)"', m[0])[1])
            hi = int(re.search(r'\bmax="(\d+)"', m[0])[1])
            return '' if lo > bound else re.sub(r'\bmax="\d+"', f'max="{min(bound, hi)}"', m[0])
        xml = re.sub(r'<col\b[^>]*/>', trim_col, xml)
        xml = re.sub(r'<dimension\b[^>]*/>', '', xml)
        out = BytesIO()
        with ZipFile(out, 'w', ZIP_DEFLATED) as dest:
            for info in source.infolist():
                dest.writestr(info, xml.encode('utf-8') if info.filename == name else source.read(info.filename))
    Path(path).write_bytes(out.getvalue())
    report.add('EMPTY_FORMAT_TRIMMED', f'工作副本移除{get_column_letter(bound + 1)}列以外的远端空白格式及{len(ranges)-len(kept)}个空合并；正文未改', level='info')


class SourceError(Exception):
    pass


class Report:
    def __init__(self, source=""):
        self.source = str(source)
        self.issues = []
        self.stages = []

    def add(self, code, message, cell="", level="warn", **details):
        self.issues.append(dict(level=level, code=code, message=message,
                                source=self.source, cell=cell, **details))

    def as_dict(self):
        return dict(source=self.source, stages=self.stages, issues=self.issues)


def is_source(path):
    p = Path(path)
    return (p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
            and not any(s in p.stem for s in ("明细", "处理后", "处理后的"))
            and not any(s.lower() == "work" or s == "源文件备份" for s in p.parts))


def scan_sources(directory="."):
    root = Path(directory)
    return sorted(p for p in root.rglob("*.xlsx") if is_source(p.relative_to(root)))


@contextmanager
def working_copy(source, report=None):
    """Exclusive creation prevents overwriting the first backup, including concurrent runs."""
    source = Path(source).resolve()
    if not is_source(source.name):
        raise SourceError(f"排除文件：{source.name}")
    backup = source.parent / "源文件备份" / source.name
    backup.parent.mkdir(exist_ok=True)
    try:
        with backup.open("xb") as dst:
            try:
                with source.open("rb") as src:
                    shutil.copyfileobj(src, dst)
            except Exception:
                dst.close()
                backup.unlink(missing_ok=True)
                raise
    except FileExistsError:
        if report:
            report.add("BACKUP_EXISTS", "备份已存在，跳过且不覆盖", level="info")
    work = source.parent / "work"
    work.mkdir(exist_ok=True)
    task_dir = Path(tempfile.mkdtemp(prefix="source-", dir=work))
    try:
        target = task_dir / source.name
        shutil.copy2(source, target)
        if report:
            report.stages.append("backup_work")
        yield target
    finally:
        shutil.rmtree(task_dir)
        try:
            work.rmdir()  # Only remove an empty work directory; never other jobs.
        except OSError:
            pass


def text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y/%m/%d")
    return str(value)


def clean_lines(value):
    lines = [s.strip() for s in value.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    return re.sub(r"\n{3,}", "\n\n", "\n".join(lines)).strip()


def width(value):
    return sum(2.2 if unicodedata.east_asian_width(ch) in "WF" else 1.0 for ch in value)


def column_map(ws):
    def mapping(r):
        headers = {c: re.sub(r"\s+", "", text(ws.cell(r, c).value)) for c in range(1, ws.max_column + 1)}
        return {k: [c for c, h in headers.items() if any(s in h for s in keys)] for k, keys in KEYS.items()}
    candidates = [(sum(bool(v) for v in mapping(r).values()), r) for r in range(1, min(ws.max_row, 20) + 1)]
    _, header = max(candidates, key=lambda x: (x[0], -x[1]))
    return header, mapping(header)


def yellow(cell):
    color = cell.fill.fgColor
    if cell.fill.patternType != "solid":
        return False
    rgb = color.rgb if color.type == "rgb" else (COLOR_INDEX[color.indexed] if color.type == "indexed" and color.indexed < len(COLOR_INDEX) else "")
    if not isinstance(rgb, str) or len(rgb) < 6:
        return False
    r, g, b = (int(rgb[-6:][i:i + 2], 16) for i in (0, 2, 4))
    return r >= 220 and g >= 200 and b <= 170


def merge_ranges(ws, report):
    ranges = [[m.min_row, m.min_col, m.max_row, m.max_col] for m in ws.merged_cells.ranges]
    for i, a in enumerate(ranges):
        for b in ranges[:i]:
            if a[0] <= b[2] and b[0] <= a[2] and a[1] <= b[3] and b[1] <= a[3]:
                report.add("MERGE_INVALID", f"合并区重叠：{a} / {b}", level="error")
                raise SourceError("合并区异常，停止导出以保护数据")
    return ranges


def detach_merges(ws, ranges):
    styles = {(r, c): copy(ws.cell(r, c)._style) for t, l, b, rr in ranges
              for r in range(t, b + 1) for c in range(l, rr + 1)}
    for t, l, b, r in ranges:
        ws.unmerge_cells(start_row=t, start_column=l, end_row=b, end_column=r)
    for (r, c), style in styles.items():
        ws.cell(r, c)._style = style


def attach_merges(ws, ranges):
    for t, l, b, r in ranges:
        if b > t or r > l:
            ws.merge_cells(start_row=t, start_column=l, end_row=b, end_column=r)


def move_dimensions(ws, row_map):
    dimensions = [(r, copy(d)) for r, d in ws.row_dimensions.items()]
    ws.row_dimensions.clear()
    for r, d in dimensions:
        n = row_map(r)
        if n is not None:
            d.index = n
            ws.row_dimensions[n] = d


def move_references(ws, row_map):
    """Move layout references, not formulas (which require explicit human review)."""
    def ref(value):
        chunks = []
        for part in str(value).split():
            prefix, bare = part.rsplit("!", 1) if "!" in part else ("", part)
            c1, r1, c2, r2 = range_boundaries(bare)
            if r1 is None or r2 is None:
                chunks.append(part)
                continue
            survivors = [n for r in range(r1, r2 + 1) if (n := row_map(r)) is not None]
            if survivors:
                chunks.append((prefix + "!" if prefix else "") + f"{get_column_letter(c1)}{min(survivors)}:{get_column_letter(c2)}{max(survivors)}")
        return " ".join(chunks)
    if ws.print_area:
        # PrintArea is an iterable in 3.0, a formatted string in 3.1.
        areas = list(ws.print_area) if not isinstance(ws.print_area, str) else str(ws.print_area).split(",")
        ws.print_area = [v for a in areas if (v := ref(a))]
    if ws.auto_filter.ref:
        ws.auto_filter.ref = ref(ws.auto_filter.ref) or None
    for dv in ws.data_validations.dataValidation:
        dv.sqref = ref(dv.sqref)
    if ws.freeze_panes:
        cell = ws[ws.freeze_panes]
        ws.freeze_panes = f"{get_column_letter(cell.column)}{row_map(cell.row) or cell.row}"
    for br in ws.row_breaks.brk:
        br.id = row_map(br.id) or max(1, (row_map(br.id - 1) or 1))
    if ws.print_title_rows:
        start, end = [int(v.replace("$", "")) for v in ws.print_title_rows.split(":")]
        ws.print_title_rows = f"{row_map(start) or start}:{row_map(end) or end}"


def parse_date(parts):
    return date(*map(int, parts))


def validate_time(value, report, cell, date_value=""):
    """Return normalized text only for a fully understood interval; never invent times."""
    original = value
    s = unicodedata.normalize("NFKC", value).replace("：", ":")
    s = re.sub(r"[—–~～]|至|到", "-", s)
    dates = list(re.finditer(DATE, s))
    clocks = list(re.finditer(CLOCK, s))
    parsed_dates = []
    try:
        parsed_dates = [parse_date(m.groups()) for m in dates]
        if not dates and date_value:
            dm = re.fullmatch(DATE, text(date_value).strip())
            if not dm:
                raise ValueError("日期列不能解析")
            parsed_dates = [parse_date(dm.groups())]
    except ValueError:
        report.add("DATE_INVALID", "日期非法，保留原值", cell, "error", value=original)
        return original, []
    if not parsed_dates:
        report.add("TIME_DATE_MISSING", "时间缺少可确定的日期，保留原值", cell, "error", value=original)
    if len(clocks) != 2:
        report.add("TIME_FORMAT", "需两个完整时刻（如08:30-17:30），未自动补造时刻", cell, "error", value=original)
        return original, parsed_dates
    times = [tuple(map(int, m.groups())) for m in clocks]
    if any(h > 24 or mi > 59 or (h == 24 and mi != 0) for h, mi in times) or times[0][0] == 24:
        report.add("TIME_INVALID", "时刻超出范围；24:00只允许作结束时间", cell, "error", value=original)
        return original, parsed_dates
    remainder = re.sub(DATE, "", re.sub(CLOCK, "", s))
    # Only whitespace and interval separators may remain; reject ambiguous multi-period text.
    if len(dates) > 2 or re.sub(r"[\s-]", "", remainder):
        report.add("TIME_FORMAT", "时间格式无法唯一解析，保留原值", cell, "error", value=original)
        return original, parsed_dates
    if dates and (dates[0].start() > clocks[0].start() or (len(dates) == 2 and not clocks[0].end() <= dates[1].start() < clocks[1].start())):
        report.add("TIME_FORMAT", "日期与时刻排列不明确，保留原值", cell, "error", value=original)
        return original, parsed_dates
    if parsed_dates:
        start = datetime.combine(parsed_dates[0], datetime.min.time()) + timedelta(hours=times[0][0], minutes=times[0][1])
        end = datetime.combine(parsed_dates[-1], datetime.min.time()) + timedelta(hours=times[1][0], minutes=times[1][1])
        if end <= start:
            report.add("TIME_ORDER", "结束时间不晚于开始时间；跨夜请明确结束日期", cell, "error", value=original)
            return original, parsed_dates
    first, second = clocks
    middle = s[first.end():second.start()]
    if not re.match(r"\s*-", middle):
        s = s[:first.end()] + "-" + s[first.end():]
    if s != original:
        report.add("TIME_FIXED", "已统一时间符号或补上两个时刻间的“-”", cell, "info", before=original, after=s)
    return s, parsed_dates


def context_range(value, year, report, label):
    full = list(re.finditer(DATE, value))
    try:
        if full:
            ds = [parse_date(m.groups()) for m in full]
            if len(ds) == 1:
                return ds[0], ds[0]
            if ds[-1] < ds[0]:
                raise ValueError()
            return ds[0], ds[-1]
        m = re.search(r"(\d{1,2})[月.](\d{1,2})日?\s*[-—–~～至]\s*(\d{1,2})[月.](\d{1,2})日?", value)
        if m and year:
            m1, d1, m2, d2 = map(int, m.groups())
            return date(year, m1, d1), date(year + (m2 < m1), m2, d2)
    except ValueError:
        report.add("DATE_CONTEXT_INVALID", f"{label}日期非法，保留原值", level="error")
    return None


def preprocess_sheet(ws, report):
    report.stages.append("preprocess")
    header, cols = column_map(ws)
    for name in ("content", "time", "measures"):
        if not cols[name]:
            report.add("MEASURES_MISSING" if name == "measures" else "COLUMN_MISSING", f"必需列未找到：{KEYS[name][0]}", level="error")
    for key, cs in cols.items():
        if len(cs) > 1 and key in ("time", "content", "date"):
            report.add("COLUMN_AMBIGUOUS", f"发现多个{KEYS[key][0]}列，逐列校验，不猜列位置")
    merges = merge_ranges(ws, report)
    deleted = set()
    for r in range(header + 1, ws.max_row + 1):
        cells = [ws.cell(r, c) for c in range(1, ws.max_column + 1)]
        risk_col = next((c for c in cols["risk"] if "作业" in text(ws.cell(header,c).value)), next(iter(cols["risk"]), None))
        risk_merge = next((m for m in merges if risk_col and m[0] <= r <= m[2] and m[1] <= risk_col <= m[3]), None)
        risk_value = text(ws.cell(risk_merge[0], risk_merge[1]).value if risk_merge else ws.cell(r, risk_col).value if risk_col else "")
        third = bool(re.search(r"(?:三|3|Ⅲ|III)\s*级|^\s*3\s*$", risk_value, re.I))
        known = third or bool(re.search(r"(?:一|二|四|五|六|[12456]|IV|VI|II|V|I)\s*级|^\s*[12456]\s*$", risk_value, re.I))
        has_yellow = any(yellow(c) for c in cells)
        if has_yellow and (third or not known):
            report.add("YELLOW_RISK_KEPT", "三级风险标黄行保留" if third else "标黄行风险级别无法确定，保留待核对", f"{ws.title}!A{r}", "info" if third else "warn")
        example = text(cells[0].value).strip() in ("例", "示例") or any(c.font.strike for c in cells) or (has_yellow and known and not third)
        filler = any(re.match(r"^填报人(?:及联系方式)?(?:\s*[:：].*)?$", text(c.value).strip()) for c in cells)
        if example or filler:
            deleted.add(r)
            report.add("ROW_DELETED", "删除示例行" if example else "删除填报人行", f"{ws.title}!A{r}", "info", original_row=r)
    if deleted:
        anchors = {(m[0], m[1]): copy(ws.cell(m[0], m[1]).value) for m in merges}
        anchor_styles = {(m[0], m[1]): copy(ws.cell(m[0], m[1])._style) for m in merges}
        detach_merges(ws, merges)
        row_map = lambda r: None if r in deleted else r - sum(d < r for d in deleted)
        move_references(ws, row_map)
        for r in sorted(deleted, reverse=True):
            ws.delete_rows(r)
        move_dimensions(ws, row_map)
        rebuilt = []
        for t, l, b, rr in merges:
            kept = [row_map(r) for r in range(t, b + 1) if r not in deleted]
            if kept:
                if t in deleted:
                    ws.cell(kept[0], l).value = anchors[t, l]
                    ws.cell(kept[0], l)._style = anchor_styles[t, l]
                    report.add("MERGE_ANCHOR_MOVED", "删除合并首行，原正文及样式已移到剩余区域首格", f"{ws.title}!{get_column_letter(l)}{t}", "info")
                rebuilt.append([kept[0], l, kept[-1], rr])
        attach_merges(ws, rebuilt)
    data_dates = []
    for r in range(header + 1, ws.max_row + 1):
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        original_row = r
        for d in sorted(deleted):
            if d <= original_row:
                original_row += 1
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell) or cell.data_type == "f":
                continue
            if isinstance(cell.value, str):
                cell.value = clean_lines(cell.value)
                if c in cols["person"] and "\n" in cell.value and re.search(r"(?:1\d{10}|0\d{2,3}-?\d{7,8})", cell.value):
                    cell.value = re.sub(r"\s*\n\s*", "", cell.value)
            if c in cols["date"] and cell.value is not None:
                try:
                    dm = re.fullmatch(DATE, text(cell.value))
                    if not dm:
                        raise ValueError()
                    data_dates.append((parse_date(dm.groups()), f"{ws.title}!{get_column_letter(c)}{original_row}"))
                except ValueError:
                    report.add("DATE_INVALID", "日期列非法或无法解析，保留原值", f"{ws.title}!{get_column_letter(c)}{original_row}", "error", value=text(cell.value))
            if c in cols["time"]:
                dv = ws.cell(r, cols["date"][0]).value if len(cols["date"]) == 1 else ""
                value, ds = validate_time(text(cell.value), report, f"{ws.title}!{get_column_letter(c)}{original_row}", dv)
                if isinstance(cell.value, str):
                    cell.value = value
                data_dates.extend((d, f"{ws.title}!{get_column_letter(c)}{original_row}") for d in ds)
    title = " ".join(text(ws.cell(r, c).value) for r in range(1, header) for c in range(1, ws.max_column + 1))
    filename = Path(report.source).stem
    years = re.findall(r"(?<!\d)(20\d{2})(?!\d)", title + " " + filename)
    year = int(years[0]) if years else (min(d.year for d, _ in data_dates) if data_dates else None)
    ranges = [(label, context_range(value, year, report, label)) for label, value in (("标题", title), ("文件名", filename))]
    if all(rng for _, rng in ranges) and ranges[0][1] != ranges[1][1]:
        report.add("DATE_CONTEXT_MISMATCH", "标题与文件名日期范围不一致", level="error")
    for label, rng in ranges:
        if rng:
            for d, cell in data_dates:
                if not rng[0] <= d <= rng[1]:
                    report.add("DATE_MISMATCH", f"日期 {d} 不在{label}范围 {rng[0]} 至 {rng[1]} 内", cell, "error")
    weeks = [re.search(r"第\s*([0-9一二三四五六七八九十百]+)\s*周", v) for v in (title, filename)]
    if all(weeks) and weeks[0][1] != weeks[1][1]:
        report.add("WEEK_MISMATCH", "标题与文件名周次不一致，输出命名采用文件名周次", level="error")
    if not any(rng for _, rng in ranges):
        report.add("DATE_CONTEXT_MISSING", "标题/文件名未找到可确定日期范围，未将业务周次强行当作ISO周次")
    if len(ws.parent.worksheets) > 1:
        report.add("OTHER_SHEETS", "仅处理第一张工作表，其余工作表保留", level="info")
    if any(c.data_type == "f" for row in ws for c in row) or ws.tables or ws.conditional_formatting:
        report.add("REFERENCE_REVIEW", "源表含公式、表格或条件格式；插删行后请核对相关引用")
    return header, cols


def auto_width(ws, header, cols, report, preserve_widths=False):
    report.stages.append("width")
    # A single <col min=12 max=14> applies to L, M and N, not only L.
    original_dimensions = list(ws.column_dimensions.values())
    existing = {c: next((d.width for d in original_dimensions if (d.min or 0) <= c <= (d.max or d.min or 0)),
                       ws.column_dimensions[get_column_letter(c)].width if get_column_letter(c) in ws.column_dimensions else 13)
                for c in range(1, ws.max_column + 1)}
    for c in range(1, ws.max_column + 1):
        longest = max((width(line) for r in range(header, ws.max_row + 1)
                       for line in text(ws.cell(r, c).value).split("\n")), default=0)
        target = max(8, min(56 if c in cols["measures"] else 48, longest + 2.4))
        dim = ws.column_dimensions[get_column_letter(c)]
        # Existing ordinary widths take precedence over the 48 auto-sizing cap.
        dim.min = dim.max = c
        dim.width = target if c in cols["measures"] and not preserve_widths else max(existing[c] or 8, target)


def auto_height(ws, cols, report, plus=MEASURES_HEIGHT_PLUS):
    if not math.isfinite(plus) or plus < 0:
        raise SourceError("管控措施行高增量必须为非负有限数字")
    report.stages.append("height")
    merges = merge_ranges(ws, report)
    needed = {}
    vertical = []
    for r in range(1, ws.max_row + 1):
        height = 15.0
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "center"
            cell.alignment = alignment
            value = text(cell.value)
            if not value:
                continue
            merge = next((m for m in merges if m[0] == r and m[1] == c), [r, c, r, c])
            col_width = sum(ws.column_dimensions[get_column_letter(cc)].width for cc in range(c, merge[3] + 1))
            size = cell.font.sz or 11
            lines = sum(max(1, math.ceil(width(s) * size / 11 / max(1, col_width - 2.4))) for s in value.split("\n"))
            h = max(18, lines * size * 1.5 + 6)
            if c in cols["measures"] or "管控措施" in value:
                h += plus
            if merge[2] > r:
                vertical.append((r, merge[2], h))
            else:
                height = max(height, h)
        needed[r] = height
    # Existing vertical blocks already have usable height in their other rows.
    # Put only the remaining shortfall on the master row, avoiding double height.
    for top, bottom, required in sorted(vertical, key=lambda m: m[1]-m[0]):
        available = sum(needed[r] for r in range(top, bottom + 1))
        needed[top] += max(0, required - available)
    for r, height in needed.items():
        ws.row_dimensions[r].height = height
        # openpyxl customHeight is a read-only computed property: setting height emits customHeight="1".
        assert ws.row_dimensions[r].customHeight
    return needed


def split_overflow(ws, needed, report):
    report.stages.append("insert")
    merges = merge_ranges(ws, report)
    detach_merges(ws, merges)
    inserted = 0
    max_col = ws.max_column
    for r in sorted(needed, reverse=True):
        required = needed[r]
        if required <= MAX_HEIGHT:
            continue
        count = math.ceil(required / MAX_HEIGHT) - 1
        if ws.max_row + count > 1048576:
            raise SourceError("插行将超出Excel最大行数")
        styles = [copy(ws.cell(r, c)._style) for c in range(1, max_col + 1)]
        covered = {c for t, l, b, rr in merges if t <= r <= b for c in range(l, rr + 1)}
        row_map = lambda old: old + count if old > r else old
        move_references(ws, row_map)
        ws.insert_rows(r + 1, count)  # 1. insert, 2. copy style, 3. merge after ALL insertions
        move_dimensions(ws, row_map)
        for new_r in range(r + 1, r + count + 1):
            for c in range(1, max_col + 1):
                ws.cell(new_r, c)._style = copy(styles[c - 1])
        for rr in range(r, r + count + 1):
            ws.row_dimensions[rr].height = required / (count + 1)
        for m in merges:
            if m[0] > r:
                m[0] += count
            if m[2] >= r:
                m[2] += count
        merges.extend([r, c, r + count, c] for c in range(1, max_col + 1) if c not in covered)
        inserted += count
        report.add("ROW_SPLIT", f"整行需求{required:.1f}磅，分摊至{count + 1}行", f"{ws.title}!A{r}", "info", inserted=count)
    attach_merges(ws, merges)
    # Set alignment on merged boundary cells too, without writing their read-only values.
    for row in ws:
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "center"
            cell.alignment = alignment
    return inserted


def format_sheet(ws, header, cols, report, plus=MEASURES_HEIGHT_PLUS, preserve_widths=True):
    auto_width(ws, header, cols, report, preserve_widths)
    needed = auto_height(ws, cols, report, plus)
    return split_overflow(ws, needed, report)


def setup_source_print(ws, header, cols, report):
    """A3 layout preserves every existing font and bold setting."""
    original_scale = ws.page_setup.scale
    print_col = ws.max_column
    if ws.print_area:
        candidates = [range_boundaries(ref)[2] for ref in re.findall(r'\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+', str(ws.print_area))]
        if candidates and max(candidates) >= max(cols['measures']):
            print_col = min(print_col, max(candidates))
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.35
    ws.print_area = f"A1:{get_column_letter(print_col)}{ws.max_row}"
    ws.print_title_rows = f"1:{header}"
    ws.row_breaks = RowBreak()
    ws.col_breaks = ColBreak()
    width_points = sum((ws.column_dimensions[get_column_letter(c)].width * 7 + 5) * 0.75 for c in range(1, print_col + 1))
    scale = min(1, max(0.1, (1190.55 - 43.2) / width_points))
    # Finished source templates deliberately use 10% with 36/48 pt type.
    # Retain their scale when it fits A3; changing it changes the printed type size.
    if original_scale and 10 <= original_scale <= 400 and width_points * original_scale / 100 <= 1147.35:
        ws.page_setup.scale = original_scale
        ws.sheet_properties.pageSetUpPr.fitToPage = False
        scale = original_scale / 100
    budget = (841.89 - 50.4 - 24) / scale - sum(ws.row_dimensions[r].height or 15 for r in range(1, header + 1))
    ranges = list(ws.merged_cells.ranges)
    used = 0
    r = header + 1
    while r <= ws.max_row:
        end = max([r] + [m.max_row for m in ranges if m.min_row <= r <= m.max_row and m.min_col <= 2 and m.max_col <= 2])
        total = sum(ws.row_dimensions[rr].height or 15 for rr in range(r, end + 1))
        if total <= budget:
            if used and used + total > budget:
                ws.row_breaks.append(Break(id=r-1)); used = 0
            used += total
        else:
            # An entire job may span pages; never deliberately cut through a measures cell.
            if used:
                ws.row_breaks.append(Break(id=r-1)); used = 0
            rr = r
            while rr <= end:
                atomic_end = max([rr] + [m.max_row for m in ranges if m.min_row <= rr <= m.max_row and any(m.min_col <= c <= m.max_col for c in cols['measures'])])
                atomic_end = min(atomic_end, end)
                height = sum(ws.row_dimensions[n].height or 15 for n in range(rr, atomic_end+1))
                if used and used + height > budget:
                    ws.row_breaks.append(Break(id=rr-1)); used = 0
                if height > budget:
                    report.add('PRINT_BLOCK_OVERSIZE', '单个管控措施合并块超过A3一页可用高度，请核对打印预览；未截断或删除正文', f'{ws.title}!A{rr}')
                used += height
                rr = atomic_end + 1
        r = end + 1
    mode = f'保留原打印缩放{original_scale}%' if ws.page_setup.scale else '宽度一页、高度不限'
    report.add('A3_PRINT', f'A3横向、{mode}；已按作业/管控措施合并块设置{len(ws.row_breaks.brk)}处分分页，原字体和加粗保留', level='info')


def output_name(source, title=""):
    match = re.search(r"第\s*([0-9一二三四五六七八九十百]+)\s*周", Path(source).stem)
    if not match:
        match = re.search(r"第\s*([0-9一二三四五六七八九十百]+)\s*周", title)
    stem = f"（第{match[1]}周）" if match else Path(source).stem
    return stem + "（处理后的源表）.xlsx"


def process_file(source, plus=MEASURES_HEIGHT_PLUS, output_dir=None, preserve_widths=True):
    source = Path(source).resolve()
    report = Report(source.name)
    target_dir = Path(output_dir) if output_dir else source.parent
    target = target_dir / output_name(source)
    report_path = target.with_suffix(".报告.json")
    wb = None
    try:
        with working_copy(source, report) as work:
            trim_inflated_copy(work, report)
            wb = load_workbook(work)
            ws = wb.worksheets[0]
            header, cols = preprocess_sheet(ws, report)
            format_sheet(ws, header, cols, report, plus, preserve_widths)
            setup_source_print(ws, header, cols, report)
            target = target_dir / output_name(source, text(ws["A1"].value))
            report_path = target.with_suffix(".报告.json")
            staged = work.parent / "output.xlsx"
            wb.save(staged)
            wb.close()
            wb = None
            target_dir.mkdir(parents=True, exist_ok=True)
            # Stage in the output directory so replace is atomic even across disks.
            with tempfile.NamedTemporaryFile(dir=target_dir, prefix=".source-", suffix=".xlsx", delete=False) as tmp:
                temporary = Path(tmp.name)
            try:
                shutil.copyfile(staged, temporary)
                temporary.replace(target)
            finally:
                temporary.unlink(missing_ok=True)
            report.stages.append("output")
    except Exception as exc:
        report.add("OUTPUT_FAILED", f"处理/输出失败：{exc}", level="error")
        raise SourceError(str(exc)) from exc
    finally:
        if wb:
            wb.close()
        report.stages.append("work_cleanup")
        try:
            report_path.write_text(json.dumps(report.as_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        except OSError as exc:
            print(f"报告写入失败：{exc}\n" + json.dumps(report.as_dict(), ensure_ascii=False))
    return target, report


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("sources", nargs="*", type=Path)
    parser.add_argument("--directory", default=".")
    parser.add_argument("--measures-plus", type=float, default=MEASURES_HEIGHT_PLUS)
    parser.add_argument("--strict-widths", action="store_true", help="将管控措施列强制收窄到56；默认保留成品宽列")
    args = parser.parse_args()
    failed = False
    for source in args.sources or scan_sources(args.directory):
        try:
            target, report = process_file(source, args.measures_plus, preserve_widths=not args.strict_widths)
            print(f"已输出 {target}；报告 {len(report.issues)} 项")
        except Exception as exc:
            failed = True
            print(f"失败 {source}: {exc}")
    return int(failed)


if __name__ == "__main__":
    raise SystemExit(main())
