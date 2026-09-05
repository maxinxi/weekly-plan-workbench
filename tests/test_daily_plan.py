"""Small synthetic tests; never commit real appointments or private workbooks."""
import base64
import copy
import io
import json
from pathlib import Path
import sys
import tempfile
import unittest
from unittest.mock import patch
from zipfile import ZipFile, ZIP_DEFLATED
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font,PatternFill
import daily_plan_server as d


def sample(kind='site'):
    w=Workbook();s=w.active;s.title='现场' if kind=='site' else '风险'
    headers=['序号','填报人','管理专业','公司名称','工作内容','关键风险点及管控措施','涉及可能发生的严重违章','当日挂接视频终端编号','检修（施工）单位','分包、外协单位','负责人姓名+联系方式','当日计划开始时间','当日计划竣工时间','到岗到位人员','同进同出人员','同进同出人员层级','地理位置','作业风险等级','电网风险等级','督查计划'] if kind=='site' else ['序号','当日计划开始时间','当日计划竣工时间','作业风险等级','电网风险等级','工作内容','负责人姓名+联系方式','成员','安全措施','危险点','工器具','平台']
    s.append(['某公司9月7日'+('现场作业计划' if kind=='site' else '作业计划（风险管控）')]);s.append(headers)
    for rr in (3,4,5):
        s.cell(rr,1,'例' if rr==3 else rr-3)
        fields={5:'更换设备'+str(rr),12:'2026-09-07 08:00',13:'2026-09-07 17:00',11:'测试负责人',3:'配电',9:'测试单位',18:'三级',19:'无',14:'到岗测试'} if kind=='site' else {6:'更换设备'+str(rr),2:'2026-09-07 08:00',3:'2026-09-07 17:00',7:'测试负责人',4:'三级',5:'无'}
        for col,v in fields.items():s.cell(rr,col,v)
        if kind=='site' and rr==4:s.cell(rr,16,'附件4-31-领导人员')
        if kind=='site' and rr==5:s.cell(rr,16,'附件5-32-管理人员')
        s.cell(rr,18 if kind=='site' else 4).fill=PatternFill('solid',fgColor='FFFF00')
    # A formatting-only continuation is not another job.
    s.cell(6,1).font=Font(size=45)
    if kind=='site':
        helper=w.create_sheet('Sheet4');helper.sheet_state='hidden'
        helper.append(['测试甲\n13000000001','测试丙\n13000000003']);helper.append(['测试乙\n13000000002'])
        dv=DataValidation(type='list',formula1='Sheet4!$A$1:$A$2');dv.add('O3');s.add_data_validation(dv)
        own=DataValidation(type='list',formula1='Sheet4!$B$1:$B$1');own.add('O5');s.add_data_validation(own)
    out=io.BytesIO();w.save(out);return out.getvalue()


class DailyTests(unittest.TestCase):
    def test_current_hidden_rules_and_valid_records(self):
        doc=d.read_document('9月7日现场.xlsx',sample(),'site')
        self.assertEqual(len(doc['records']),2)
        a,b=doc['records'];self.assertEqual(len(a['options']),2);self.assertEqual(len(b['options']),1)
        self.assertTrue(a['inherited']);self.assertFalse(b['inherited'])
        self.assertTrue(a['needsSelection']);self.assertFalse(b['needsSelection'])
        self.assertEqual(doc['hiddenSheets'],['Sheet4'])

    def test_only_blank_leader_rows_require_selection(self):
        docs={k:d.read_document('9月7日'+k+'.xlsx',sample(k),k) for k in ('site','risk')}
        sess=d.Session(tempfile.gettempdir());sess.documents=copy.deepcopy(docs)
        view=sess.view();self.assertEqual(view['selection'],{'required':1,'completed':0});self.assertFalse(view['ready'])
        self.assertFalse(any(i['code']=='DROPDOWN' and i['row']==5 for i in view['issues']))
        with self.assertRaisesRegex(ValueError,'不提供同进同出选择'):
            sess.update([dict(kind='site',row=5,field='same',value='不应写入')])
        sess.update([dict(kind='site',row=4,field='same',value='测试甲')])
        self.assertTrue(sess.view()['ready'])

    def test_patch_preserves_package_and_clones_only_missing_rule(self):
        data=sample();original=d.read_document('9月7日现场.xlsx',data,'site');doc=copy.deepcopy(original)
        doc['records'][0]['same']='上午：测试甲\n13000000001\n下午：测试乙\n13000000002'
        output=d.patch_package(data,doc,original)
        with ZipFile(io.BytesIO(data)) as a,ZipFile(io.BytesIO(output)) as b:
            for name in a.namelist():
                if name!='xl/worksheets/sheet1.xml':self.assertEqual(a.read(name),b.read(name),name)
        w=load_workbook(io.BytesIO(output));self.assertEqual(w.worksheets[0]['O4'].value,doc['records'][0]['same'])
        self.assertTrue(any('O4' in str(v.sqref) for v in w.worksheets[0].data_validations.dataValidation))
        self.assertEqual(w.worksheets[0]['R4'].fill.fgColor.rgb[-6:],'FFFF00')

    def test_x14_validation_read_and_preserve(self):
        data=sample();out=io.BytesIO()
        with ZipFile(io.BytesIO(data)) as a,ZipFile(out,'w',ZIP_DEFLATED) as b:
            for item in a.infolist():
                v=a.read(item.filename)
                if item.filename=='xl/worksheets/sheet1.xml':
                    ext='<extLst><ext uri="test"><x14:dataValidations xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main"><x14:dataValidation type="list"><x14:formula1><xm:f>Sheet4!$B$1:$B$1</xm:f></x14:formula1><xm:sqref>O4</xm:sqref></x14:dataValidation></x14:dataValidations></ext></extLst>'
                    v=v.decode().replace('</worksheet>',ext+'</worksheet>').encode()
                b.writestr(item,v)
        doc=d.read_document('9月7日现场.xlsx',out.getvalue(),'site');self.assertEqual(doc['records'][0]['options'],['测试丙\n13000000003'])
        changed=copy.deepcopy(doc);changed['records'][0]['same']='测试丙'
        result=d.patch_package(out.getvalue(),changed,doc)
        with ZipFile(io.BytesIO(result)) as z:self.assertIn(b'x14:dataValidation',z.read('xl/worksheets/sheet1.xml'))

    def test_relative_dropdown_reference_tracks_destination(self):
        w=load_workbook(io.BytesIO(sample()));s=w.worksheets[0]
        s.data_validations.dataValidation=[]
        dv=DataValidation(type='list',formula1='INDIRECT(P3)');dv.add('O3:O5');s.add_data_validation(dv)
        b=io.BytesIO();w.save(b)
        doc=d.read_document('9月7日现场.xlsx',b.getvalue(),'site')
        self.assertEqual([r['formula'] for r in doc['records']],['INDIRECT(P4)','INDIRECT(P5)'])
        self.assertTrue(all(r['formulaError'] for r in doc['records']))

    def test_only_serial_empty_row_removed_but_continuation_preserved(self):
        w=load_workbook(io.BytesIO(sample('risk')));s=w.worksheets[0];s['A6']=3
        b=io.BytesIO();w.save(b);self.assertEqual(d.serial_only_rows(b.getvalue(),'risk'),[6])
        s['I6']='仍有安全措施';b=io.BytesIO();w.save(b);self.assertEqual(d.serial_only_rows(b.getvalue(),'risk'),[])
        s['I6']=None;s.merge_cells('F5:F6');b=io.BytesIO();w.save(b)
        self.assertEqual(d.serial_only_rows(b.getvalue(),'risk'),[],'do not delete a merged continuation that holds printed work content')

    def test_date_time_and_cross_file_warnings(self):
        docs={k:d.read_document('9月7日'+k+'.xlsx',sample(k),k) for k in ('site','risk')}
        self.assertFalse(any(i['code'] in ('COUNT','CROSS_TIME','UNMATCHED') for i in d.audit(docs)))
        docs['risk']['records'][0]['start']='2026-09-07 09:00'
        self.assertTrue(any(i['code']=='CROSS_TIME' and i['suggestion']=='2026-09-07 08:00' for i in d.audit(docs)))
        for value in ('2026-02-30 08:00','08:30','2026-09-07 25:00'):
            self.assertIsNone(d.parse_time(value)[0])
        docs['site']['records'][0]['end']='2026-09-07 06:00';self.assertIn('TIME_ORDER',[i['code'] for i in d.audit(docs)])
        docs['site']['title']='某公司9月8日现场作业计划';self.assertIn('TITLE_DATE',[i['code'] for i in d.audit(docs)])
        docs['risk']['records'].pop();self.assertIn('COUNT',[i['code'] for i in d.audit(docs)])

    def test_cleaned_composite_matching_ignores_order(self):
        docs={k:d.read_document('9月7日'+k+'.xlsx',sample(k),k) for k in ('site','risk')}
        docs['risk']['records'].reverse()
        self.assertFalse(any(i['code'] in ('COUNT','UNMATCHED','CROSS_TIME') for i in d.audit(docs)))
        docs['risk']['records'][0]['work']='措辞稍有不同'
        codes=[i['code'] for i in d.audit(docs)]
        self.assertIn('CROSS_WORK',codes);self.assertNotIn('UNMATCHED',codes)
        self.assertEqual(docs['site']['cleanup']['exampleRowsRemoved'],[3])
        self.assertIn(6,docs['site']['cleanup']['blankRowsRemoved'])

    def test_backup_not_overwritten_and_explicit_edit(self):
        with tempfile.TemporaryDirectory() as t:
            p=Path(t)/'源文件备份'/'input.xlsx';d.backup_once(p,b'original');d.backup_once(p,b'changed');self.assertEqual(p.read_bytes(),b'original')
            sess=d.Session(t);files=[dict(name='9月7日'+('现场' if k=='site' else '风险管控')+'.xlsx',data=base64.b64encode(sample(k)).decode()) for k in ('site','risk')]
            before=sess.load(files);self.assertFalse(before['ready'])
            frozen=dict(sess.inputs);sess.update([dict(kind='site',row=4,field='same',value='上午：测试甲\n下午：测试乙')]);self.assertEqual(sess.inputs,frozen)
            self.assertEqual(sess.documents['risk']['records'][0]['owner'],'测试负责人')
            self.assertEqual(len(sess.changes),1)

    def test_failed_generation_cleans_work(self):
        with tempfile.TemporaryDirectory() as t:
            sess=d.Session(t);sess.load([dict(name='9月7日'+('现场' if k=='site' else '风险管控')+'.xlsx',data=base64.b64encode(sample(k)).decode()) for k in ('site','risk')])
            for r in sess.documents['site']['records']:r['same']='测试人员'
            from contextlib import nullcontext
            with patch.object(d,'a3_printer',return_value=nullcontext()),patch.object(d.legacy(),'preprocess_excel',side_effect=RuntimeError('test output failure')):
                sess.generate({'acknowledged':[i['id'] for i in sess.view()['issues']]})
            self.assertIn('test output failure',sess.status['error']);self.assertFalse(list(Path(t).glob('work-*')))
            self.assertFalse(sess.result)

    def test_source_directory_output_and_atomic_rollback(self):
        with tempfile.TemporaryDirectory() as t:
            base=Path(t);app=base/'app';app.mkdir()
            paths=[]
            for k in ('site','risk'):
                p=base/('9月7日'+('现场' if k=='site' else '风险管控')+'.xlsx');p.write_bytes(sample(k));paths.append(p)
            sess=d.Session(app);view=sess.load_paths(paths)
            self.assertEqual(view['outputLocation'],str(base/'（处理后）'))
            for p in paths:self.assertEqual(p.read_bytes(),(base/'源文件备份'/p.name).read_bytes())
            source=base/'staged';source.mkdir();(source/'a.xlsx').write_bytes(b'new-a');(source/'b.pdf').write_bytes(b'new-b')
            dest=base/'（处理后）';dest.mkdir();(dest/'a.xlsx').write_bytes(b'old-a')
            d.publish_output(source,dest);self.assertEqual((dest/'a.xlsx').read_bytes(),b'new-a')
            (source/'a.xlsx').write_bytes(b'next-a')
            real_replace=d.os.replace
            def locked(a,b):
                if Path(b).name=='b.pdf':raise PermissionError('locked')
                return real_replace(a,b)
            with patch.object(d.os,'replace',side_effect=locked),self.assertRaises(PermissionError):d.publish_output(source,dest)
            self.assertEqual((dest/'a.xlsx').read_bytes(),b'new-a')
            self.assertFalse(list(base.glob('work-*')))

    def test_summary_modes_use_original_rules(self):
        doc=d.read_document('9月7日现场.xlsx',sample(),'site')
        a=d.summaries(doc);self.assertIn('【缩写版】',a['a']);self.assertIn('【完整版】',a['a']);self.assertEqual(a['b'],'')
        with self.assertRaises(ValueError):d.summaries(doc,True)
        for r in doc['records']:r['same']='测试甲'
        self.assertIn('工作计划2项',d.summaries(doc,True)['b'])

    def test_missing_column_is_not_guessed(self):
        data=sample();w=load_workbook(io.BytesIO(data));w.worksheets[0]['O2']='人员层级';b=io.BytesIO();w.save(b)
        with self.assertRaisesRegex(ValueError,'同进同出人员'):d.read_document('现场.xlsx',b.getvalue(),'site')
        for name in ('../源表.xlsx','~$源表.xlsx','C:\\源表.xlsx'):
            with self.assertRaises(ValueError):d.safe_name(name)


if __name__=='__main__':unittest.main()
