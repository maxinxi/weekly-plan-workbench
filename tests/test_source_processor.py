import importlib.util
import json
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from source_processor import (Report, SourceError, working_copy, process_file, preprocess_sheet,
                              auto_width, auto_height, split_overflow, output_name, scan_sources,
                              validate_time, format_sheet, trim_inflated_copy, detail_name, output_folder)


def fixture(path):
    wb=Workbook(); ws=wb.active; ws.title='表1'
    ws.append(['2026年9月1日-2026年9月7日 第36周'])
    ws.merge_cells('A1:H1')
    ws.append(['序号','作业内容','负责人','计划时间','管控措施','风险等级','横向区域',''])
    ws.append(['示例','删除我','', '', '示例内容'])
    ws.append(['1','保留甲','张三\r\n13800000000','2026/9/1 08:30\n17:30','措施\r\n\r\n\r\n测试',''])
    ws.append(['','填报人及联系方式：测试','','','',''])
    ws.append(['2','保留乙','李四','2026/9/2 08:30-17:30','安全措施。'*900,'备注\n第二行','横向合并'])
    ws.merge_cells('G6:H6')
    ws.append(['3','保留下方','王五','2026/9/3 08:30-17:30','短措施','下方边框'])
    ws.merge_cells('G7:H7')
    ws.append(['例','删除我2'])
    ws.append(['4','标黄删除'])
    ws['B9'].fill=PatternFill('solid',fgColor='FFFFFF00')
    ws['F9']='四级'
    ws.append(['5','删除线删除'])
    ws['B10'].font=Font(strike=True)
    ws.append(['6','保留末尾','人员','2026/9/4 08:30-17:30','末尾措施'])
    ws.append([None]*8); ws.row_dimensions[12].height=100
    for row in ws:
        for c in row:
            if c.row!=10:
                c.font=Font(name='宋体',size=11)
            c.border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ws.column_dimensions['B'].width=62
    ws.column_dimensions['E'].width=255
    ws.print_area='A1:H12'; ws.print_title_rows='1:2'; ws.freeze_panes='A3'
    wb.save(path); wb.close()


class SourceTests(unittest.TestCase):
    def setUp(self):
        self.temp=tempfile.TemporaryDirectory(); self.root=Path(self.temp.name)
        self.src=self.root/'表1（第36周）.xlsx'; fixture(self.src)
    def tearDown(self): self.temp.cleanup()

    def test_backup_readonly_cleanup_and_pipeline(self):
        original=self.src.read_bytes()
        target,report=process_file(self.src, preserve_widths=False)
        self.assertEqual(target,self.root/'（处理后）'/self.src.name)
        self.assertEqual(original,self.src.read_bytes())
        backup=self.root/'源文件备份'/self.src.name
        self.assertEqual(original,backup.read_bytes())
        self.assertFalse((self.root/'work').exists())
        self.assertEqual(report.stages[:5],['backup_work','preprocess','width','height','insert'])
        wb=load_workbook(target);ws=wb.worksheets[0]
        values=[str(c.value) for row in ws for c in row if c.value is not None]
        self.assertFalse(any('删除我' in v or v.startswith('填报人') for v in values))
        self.assertIn('张三13800000000',values)
        self.assertIn('2026/9/1 08:30-\n17:30',values)
        self.assertEqual(ws.column_dimensions['E'].width,56)
        self.assertEqual(ws.column_dimensions['B'].width,62)
        self.assertGreater(len(ws.merged_cells.ranges),3)
        for r in range(1,ws.max_row+1):
            self.assertLessEqual(ws.row_dimensions[r].height,409.5)
            self.assertTrue(ws.row_dimensions[r].customHeight)
        for row in ws:
            for c in row:
                if c.value is not None:
                    self.assertTrue(c.alignment.wrap_text)
                    self.assertEqual(c.alignment.vertical,'center')
        row=next(r for r in range(1,ws.max_row+1) if ws.cell(r,2).value=='保留乙')
        m=next(m for m in ws.merged_cells.ranges if m.min_row==row and m.min_col==7)
        self.assertEqual(m.max_col,8); self.assertGreater(m.max_row,row)
        self.assertEqual(ws.cell(m.max_row+1,2).value,'保留下方')
        for c in range(1,9):
            self.assertTrue(any(m.min_row<=row and m.max_row>row and m.min_col<=c<=m.max_col for m in ws.merged_cells.ranges))
        wb.close()
        with ZipFile(target) as z:
            xml=ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
            ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            self.assertTrue(all(r.attrib.get('customHeight')=='1' for r in xml.findall('m:sheetData/m:row',ns)))
        wb=load_workbook(self.src);wb.active['B4']='修改后的输入';wb.save(self.src);wb.close()
        process_file(self.src)
        self.assertEqual(original,backup.read_bytes())

    def test_work_cleanup_failure(self):
        with self.assertRaises(RuntimeError):
            with working_copy(self.src) as path:
                self.assertTrue(path.exists()); raise RuntimeError('test')
        self.assertFalse((self.root/'work').exists())
        with patch('source_processor.format_sheet',side_effect=RuntimeError('output failure')):
            with self.assertRaises(SourceError): process_file(self.src)
        self.assertFalse((self.root/'work').exists())
        report=json.loads(next((self.root/'（处理后）').glob('*.报告.json')).read_text('utf-8'))
        self.assertIn('OUTPUT_FAILED',[i['code'] for i in report['issues']])

    def test_small_rows_and_whole_row_trigger(self):
        wb=Workbook(); ws=wb.active
        ws.append(['序号','作业内容','时间','管控措施','备注'])
        ws.append([1,'任务','2026/9/1 08:30-17:30','短','长备注\n'*50])
        ws.append([2,'结束','2026/9/1 08:30-17:30','短'])
        log=Report(); header,cols=preprocess_sheet(ws,log)
        auto_width(ws,header,cols,log); heights=auto_height(ws,cols,log)
        self.assertGreater(heights[2],409.5); self.assertLess(heights[3],409.5)
        split_overflow(ws,heights,log)
        self.assertGreater(ws.max_row,3)
        self.assertTrue(any(m.min_col==4 and m.min_row==2 and m.max_row>2 for m in ws.merged_cells.ranges))

    def test_existing_vertical_merge(self):
        wb=Workbook();ws=wb.active
        ws.append(['序号','作业内容','时间','管控措施'])
        ws.append([1,'共用项目','2026/9/1 08:30-17:30','很长\n'*70]);ws.append([None,None,None,'子项保留'])
        ws.append([2,'下一个','2026/9/2 08:30-17:30','保留'])
        for c in range(1,4):ws.merge_cells(start_row=2,start_column=c,end_row=3,end_column=c)
        log=Report();h,cols=preprocess_sheet(ws,log);format_sheet(ws,h,cols,log)
        self.assertIn('子项保留',[ws.cell(r,4).value for r in range(1,ws.max_row+1)])
        self.assertIn('下一个',[ws.cell(r,2).value for r in range(1,ws.max_row+1)])

    def test_preserve_finished_width_font_and_merge_height(self):
        wb=Workbook();ws=wb.active
        ws.append(['序号','作业内容','计划时间','管控措施','作业风险等级'])
        ws.append([1,'保持48磅粗体','2026/9/1 08:30-17:30','短措施','三级'])
        ws.append([None,None,None,'第二条措施'])
        for c in [1,2,3,5]: ws.merge_cells(start_row=2,start_column=c,end_row=3,end_column=c)
        ws['B2'].font=Font(name='宋体',size=48,bold=True)
        ws['E2'].fill=PatternFill('solid',fgColor='FFFFFF00')
        ws.column_dimensions['B'].width=234; ws.column_dimensions['D'].width=255
        log=Report();header,cols=preprocess_sheet(ws,log)
        format_sheet(ws,header,cols,log)
        self.assertEqual(ws['B2'].font.sz,48);self.assertTrue(ws['B2'].font.bold)
        self.assertEqual(ws.column_dimensions['D'].width,255)
        self.assertEqual(ws.max_row,3)
        self.assertLess(ws.row_dimensions[2].height+ws.row_dimensions[3].height,110)
        self.assertEqual(ws['E2'].fill.fgColor.rgb,'FFFFFF00')

    def test_deleted_merge_anchor_keeps_style(self):
        wb=Workbook();ws=wb.active
        ws.append(['序号','作业内容','计划时间','管控措施','作业风险等级'])
        ws.append([1,'原正文','2026/9/1 08:30-17:30','删除措施','四级'])
        ws.append([None,None,None,'保留措施'])
        ws['B2'].font=Font(name='宋体',size=48,bold=True)
        ws['E2'].fill=PatternFill('solid',fgColor='FFFFFF00')
        for c in [1,2,3,5]: ws.merge_cells(start_row=2,start_column=c,end_row=3,end_column=c)
        log=Report();preprocess_sheet(ws,log)
        self.assertEqual(ws['B2'].value,'原正文');self.assertEqual(ws['B2'].font.sz,48)
        self.assertTrue(ws['B2'].font.bold);self.assertEqual(ws['D2'].value,'保留措施')

    def test_far_blank_merges_do_not_expand_print_area(self):
        wb=load_workbook(self.src)
        wb.active.merge_cells('XFC4:XFC5');wb.active['XFC4'].border=Border(bottom=Side(style='thin'))
        wb.save(self.src);wb.close();before=self.src.read_bytes()
        with working_copy(self.src) as work:
            log=Report();trim_inflated_copy(work,log)
            wb=load_workbook(work)
            self.assertEqual(wb.active.max_column,8)
            self.assertEqual(wb.active['G6'].value,'横向合并')
            self.assertTrue(any(i['code']=='EMPTY_FORMAT_TRIMMED' for i in log.issues));wb.close()
        self.assertEqual(self.src.read_bytes(),before)

    def test_dates_errors_and_no_guesses(self):
        cases=[('2026/2/30 08:30-17:30','DATE_INVALID'),('08:30-17:30','TIME_DATE_MISSING'),
               ('2026/9/1 25:00-26:00','TIME_INVALID'),('2026/9/1 17:30-08:30','TIME_ORDER'),
               ('2026/9/1 上午开工','TIME_FORMAT'),('2026/9/1 08:30-17:30 或 18:00','TIME_FORMAT')]
        for value,code in cases:
            log=Report();fixed,_=validate_time(value,log,'J4')
            self.assertEqual(value,fixed); self.assertIn(code,[i['code'] for i in log.issues])
        log=Report();fixed,_=validate_time('2026/9/1 08：30 17：30',log,'J4')
        self.assertEqual(fixed,'2026/9/1 08:30- 17:30')
        wb=load_workbook(self.src);wb.active['D4']='2026/10/1 08:30-17:30'
        log=Report(self.src.name);preprocess_sheet(wb.active,log)
        self.assertIn('DATE_MISMATCH',[i['code'] for i in log.issues]);wb.close()

    def test_yellow_third_level_is_kept(self):
        wb=Workbook();ws=wb.active
        ws.append(['序号','作业内容','计划时间','管控措施','作业风险等级'])
        for i,risk in enumerate(['三级','3级','四级','五级','未知'],2):
            ws.append([i,'测试'+risk,'2026/9/1 08:30-17:30','措施',risk])
            ws.cell(i,2).fill=PatternFill('solid',fgColor='FFFFFF00')
        log=Report();preprocess_sheet(ws,log)
        self.assertEqual([ws.cell(r,5).value for r in range(2,ws.max_row+1)],['三级','3级','未知'])
        self.assertEqual(sum(i['code']=='ROW_DELETED' for i in log.issues),2)

    def test_missing_columns_and_merge_error(self):
        wb=Workbook();ws=wb.active;ws.append(['序号','备注']);ws.append([1,'值'])
        log=Report();preprocess_sheet(ws,log)
        self.assertIn('MEASURES_MISSING',[i['code'] for i in log.issues])
        self.assertIn('COLUMN_MISSING',[i['code'] for i in log.issues])
        ws.merge_cells('A4:B5');ws.merge_cells('B5:C6')
        with self.assertRaises(SourceError):preprocess_sheet(ws,Report())

    def test_names_scanning(self):
        self.assertEqual(output_name('表1（第36周）.xlsx'),'表1（第36周）.xlsx')
        self.assertEqual(output_name('源表.xlsx'),'源表.xlsx')
        self.assertEqual(output_name('源表.xlsx','第九周'),'源表.xlsx')
        self.assertEqual(detail_name('表1：每周重点作业计划（第36周）.xlsx'),'（第36周）（周计划明细）.xlsx')
        self.assertEqual(detail_name('表1(第 37 周).xlsx'),'（第37周）（周计划明细）.xlsx')
        self.assertEqual(detail_name('表1.xlsx','第九周'),'（第九周）（周计划明细）.xlsx')
        self.assertEqual(detail_name('表1.xlsx'),'周计划明细.xlsx')
        for name in ['~$表1.xlsx','周计划明细.xlsx','表1（处理后的源表）.xlsx']:(self.root/name).touch()
        (self.root/'work').mkdir();(self.root/'work'/'表1.xlsx').touch()
        processed=self.root/'（处理后）';processed.mkdir();(processed/self.src.name).touch()
        self.assertEqual(output_folder(processed),processed)
        self.assertEqual(scan_sources(self.root),[self.src])

    def test_output_never_overwrites_same_name_original(self):
        target,_=process_file(self.src)
        before=target.read_bytes()
        with self.assertRaisesRegex(SourceError,'原件路径相同'):
            process_file(target)
        self.assertEqual(before,target.read_bytes())

    def test_detail_saved_beside_source_in_processed_folder(self):
        from datetime import datetime
        spec=importlib.util.spec_from_file_location('weekly_export',Path(__file__).resolve().parents[1]/'weekly-plan-export.py')
        module=importlib.util.module_from_spec(spec);spec.loader.exec_module(module)
        start=datetime(2026,8,31);end=datetime(2026,9,6)
        with patch.object(module,'safe_print'):
            module.generate_excel_output_v5({},start,end,str(self.src))
        expected=self.root/'（处理后）'/'（第36周）（周计划明细）.xlsx'
        self.assertTrue(expected.exists())
        wb=load_workbook(expected);self.assertEqual(wb.active.title,'周计划明细');wb.close()


if __name__=='__main__':unittest.main()
