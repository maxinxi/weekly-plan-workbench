"""Independent openpyxl read-back verifies the browser file matches the Python source layout."""
from pathlib import Path
import sys
import tempfile
import shutil
from zipfile import ZipFile
from openpyxl import load_workbook
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from source_processor import process_file

source=Path(sys.argv[1]); browser=Path(sys.argv[2])
with tempfile.TemporaryDirectory() as temp:
    copied=Path(temp)/source.name;shutil.copy2(source,copied)
    output,_=process_file(copied)
    py=load_workbook(output);js=load_workbook(browser)
    p,j=py.worksheets[0],js.worksheets[0]
    assert (p.max_row,p.max_column)==(j.max_row,j.max_column),((p.max_row,p.max_column),(j.max_row,j.max_column))
    assert {str(m) for m in p.merged_cells.ranges}=={str(m) for m in j.merged_cells.ranges},'merge mismatch'
    for row in p:
        for c in row:
            other=j[c.coordinate]
            assert c.value==other.value,(c.coordinate,c.value,other.value)
            if c.value is not None:
                assert c.font.sz==other.font.sz,(c.coordinate,'font')
                assert c.font.bold==other.font.bold,(c.coordinate,'bold')
                assert c.alignment.wrap_text and other.alignment.wrap_text
                assert other.alignment.vertical=='center'
    for r in range(1,p.max_row+1):
        assert abs(p.row_dimensions[r].height-j.row_dimensions[r].height)<0.05,(r,p.row_dimensions[r].height,j.row_dimensions[r].height)
        assert j.row_dimensions[r].height<=409.5
        assert j.row_dimensions[r].customHeight
    for c in range(1,p.max_column+1):
        from openpyxl.utils import get_column_letter
        key=get_column_letter(c)
        def effective_width(sheet,col):
            return next((d.width for d in sheet.column_dimensions.values() if d.min<=col<=d.max),13)
        assert abs(effective_width(p,c)-effective_width(j,c))<0.01,(key,'width',effective_width(p,c),effective_width(j,c))
    pp,jp=p,j
    assert str(jp.page_setup.paperSize)=='8' and jp.page_setup.orientation=='landscape'
    assert jp.page_setup.fitToWidth==1 and jp.page_setup.fitToHeight==0
    assert [tuple(c.value for c in row) for row in pp]==[tuple(c.value for c in row) for row in jp],'print content mismatch'
    assert all(jp.row_dimensions[r].height<=409.5 for r in range(1,jp.max_row+1))
    assert len(js.worksheets)==len(py.worksheets)
    assert [b.id for b in p.row_breaks.brk]==[b.id for b in j.row_breaks.brk],'page breaks mismatch'
    for b in j.row_breaks.brk:
        assert not any(m.min_row<=b.id<m.max_row and m.min_col<=17<=m.max_col for m in j.merged_cells.ranges),('break crosses measures',b.id)
    print(f'Python/browser parity passed: {p.max_row} source rows, A3 layout, fonts and full content')
    py.close();js.close()
with ZipFile(browser) as z:
    from xml.etree import ElementTree as ET
    for name in z.namelist():
        if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
            props=ET.fromstring(z.read(name)).find('{*}sheetPr')
            tags=[] if props is None else [el.tag.split('}')[-1] for el in props]
            if 'outlinePr' in tags and 'pageSetUpPr' in tags:
                assert tags.index('outlinePr')<tags.index('pageSetUpPr'),'desktop Excel requires CT_SheetPr order'
zip_path=browser.parent/'backup.zip'
if zip_path.exists():
    with ZipFile(zip_path) as z:
        assert z.read('源文件备份/'+source.name)==source.read_bytes()
    print('Backup ZIP is byte-for-byte identical to original')
