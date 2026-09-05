"""Verify the ZIP downloaded by the actual offline HTML, without browser libraries."""
from pathlib import Path
from io import BytesIO
from zipfile import ZipFile
import sys
from openpyxl import load_workbook
sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from source_processor import detail_name

archive=Path(sys.argv[1]);source=Path(sys.argv[2])
with ZipFile(archive) as z:
    names=[n for n in z.namelist() if not n.endswith('/')]
    expected={'（处理后）/'+source.name,'（处理后）/'+detail_name(source),
              '（处理后）/'+source.with_suffix('.报告.json').name}
    assert set(names)==expected,(names,expected)
    for name in names:
        if name.endswith('.xlsx'):
            wb=load_workbook(BytesIO(z.read(name)))
            assert wb.worksheets[0].max_row>2
            wb.close()
print('Offline ZIP verified: original source filename, week-only detail filename, report, all inside （处理后）.')
