# -*- coding: utf-8 -*-
"""
周计划工具（V5功能移植版 - 管控措施行高可调版 - 终极修复版）
功能：
1. 周计划明细 Excel 输出（基于V5的优化逻辑）
2. 三四五级风险汇总 TXT 输出
3. 周计划项目汇总文本输出
4. 处理后的源表导出（管控措施超长全列贯通插行分摊）
行高：
- 含「管控措施」的内容格：估算高度 + MEASURES_HEIGHT_PLUS（默认 +1 磅）
- 只改文件顶部的 MEASURES_HEIGHT_PLUS：1.0=+1，2.0=+2，0=不加
- 已取消原脚本「全表统一 -1」
- 超长内容（>409.5磅）自动整行插行续行并平摊行高
"""
import os
import re
import glob
import sys
import math
import traceback
from datetime import datetime, timedelta
from collections import OrderedDict
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break, RowBreak
import warnings
warnings.filterwarnings('ignore')
# -------------------- 环境初始化 --------------------
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
ERROR_LOG_PATH = "周计划导出错误日志.txt"
_error_log_inited = False
_error_count = 0
# Excel 单行最大行高（磅），超过此值需要插行续行
EXCEL_MAX_ROW_HEIGHT = 409.5
# A3 横向可打印高度估算（磅）：297mm / 25.4 * 72 ≈ 842，减页边距后约 820
A3_LANDSCAPE_PRINTABLE_HEIGHT = 820.0
PAGE_HEIGHT_SAFETY_FACTOR = 0.96
# ============================================================
# ★★★ 行高微调：只改这一处 ★★★
# 含「管控措施」的内容单元格 = 估算行高 + 此值（磅）
#   1.0 → +1     2.0 → +2     0 → 不加
# ============================================================
MEASURES_HEIGHT_PLUS = 1.0
MEASURES_SRC_HEIGHT_PLUS = 0.0
# 单次预处理缓存，防止主流程重复清洗源文件
_processed_source_cache = set()
# -------------------- 工具函数 --------------------
def safe_print(msg, error=False):
    prefix = "❌ " if error else ""
    try:
        print(f"{prefix}{msg}")
    except Exception:
        clean = re.sub(r'[\x1b\[0-9;]*[mGKHF]', '', msg)
        print(clean.encode('gbk', errors='ignore').decode('gbk', errors='ignore'))
def _ensure_error_log():
    global _error_log_inited
    if _error_log_inited:
        return
    _error_log_inited = True
    header = f"\n{'='*60}\n"
    header += f"❌ 错误日志 | 启动时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
    header += f"💻 环境: Python {sys.version.split()[0]} | 目录: {os.getcwd()}\n"
    header += f"{'='*60}\n"
    try:
        with open(ERROR_LOG_PATH, 'w', encoding='utf-8') as f:
            f.write(header)
    except Exception:
        pass
def log_error(file_name, error_type, error_msg, traceback_str=""):
    global _error_count
    _error_count += 1
    _ensure_error_log()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_entry = f"\n[{timestamp}] 源文件: {file_name}\n"
    log_entry += f"  🚨 类型: {error_type}\n"
    log_entry += f"  💡 描述: {error_msg}\n"
    if traceback_str:
        log_entry += f"  📋 堆栈:\n{traceback_str[:500]}\n"
    log_entry += "-" * 60
    try:
        with open(ERROR_LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(log_entry)
    except Exception:
        print(f"⚠️  日志写入失败！请记录: {error_msg}")
def unmerge_and_fill(ws):
    merged = list(ws.merged_cells.ranges)
    if not merged:
        return ws
    fill_data = []
    for rng in merged:
        min_row, min_col, max_row, max_col = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        try:
            val = ws.cell(min_row, min_col).value
        except Exception:
            val = None
        fill_data.append((min_row, min_col, max_row, max_col, val))
    for rng in merged:
        try:
            ws.unmerge_cells(str(rng))
        except Exception:
            pass
    for min_row, min_col, max_row, max_col, val in fill_data:
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                try:
                    ws.cell(r, c).value = val
                except Exception:
                    pass
    return ws
def normalize_content(s):
    if pd.isna(s): return ""
    return re.sub(r'\s+', ' ', str(s)).strip()
def normalize_clock_text(value):
    text = str(value).strip()
    m = re.fullmatch(r'(\d{1,2}):(\d{2})', text)
    if not m:
        return text
    hour, minute = map(int, m.groups())
    if not (0 <= minute <= 59 and 0 <= hour <= 24):
        return text
    if hour == 24 and minute != 0:
        return text
    return f"{hour:02d}:{minute:02d}"
def canonicalize_time(s):
    if pd.isna(s):
        return ""
    text = str(s).strip()
    pattern = r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s*(\d{1,2}:\d{2})\s*[-–]\s*(\d{4}[/-]\d{1,2}[/-]\d{1,2})?\s*(\d{1,2}:\d{2})?'
    m = re.search(pattern, text)
    if not m:
        pattern_no_dash = r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s+(\d{1,2}:\d{2})\s+(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s+(\d{1,2}:\d{2})'
        m = re.search(pattern_no_dash, text)
    if not m:
        return text
    d1, t1, d2, t2 = m.groups()
    d1 = d1.replace('-', '/')
    d2 = (d2 or d1).replace('-', '/')
    try:
        d1_norm = datetime.strptime(d1, "%Y/%m/%d").strftime("%Y/%m/%d")
        d2_norm = datetime.strptime(d2, "%Y/%m/%d").strftime("%Y/%m/%d")
    except ValueError:
        return text
    t1_norm = normalize_clock_text(t1)
    t2_default = "18:00" if d1_norm != d2_norm else "17:30"
    t2_norm = normalize_clock_text(t2 or t2_default)
    return f"{d1_norm}|{t1_norm}|{d2_norm}|{t2_norm}"
def extract_daily_time_segments(time_str):
    if pd.isna(time_str):
        return {}
    s = str(time_str).strip()
    pat_a = r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s*(\d{1,2}:\d{2})\s*[-–]\s*(\d{4}[/-]\d{1,2}[/-]\d{1,2})?\s*(\d{1,2}:\d{2})?'
    m = re.search(pat_a, s)
    if not m:
        pat_a2 = r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s+(\d{1,2}:\d{2})\s+(\d{4}[/-]\d{1,2}[/-]\d{1,2})\s+(\d{1,2}:\d{2})'
        m = re.search(pat_a2, s)
    if m:
        d1, t1, d2, t2 = m.groups()
        d1 = d1.replace('-', '/')
        d2 = (d2 or d1).replace('-', '/')
        try:
            start_date = datetime.strptime(d1, "%Y/%m/%d")
            end_date = datetime.strptime(d2, "%Y/%m/%d")
        except ValueError:
            return {}
        if end_date < start_date:
            return {}
        t1 = normalize_clock_text(t1)
        t2 = normalize_clock_text(t2) if t2 else None
    else:
        pat_b = r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})'
        m = re.search(pat_b, s)
        if not m:
            return {}
        d1 = m.group(1).replace('-', '/')
        try:
            start_date = datetime.strptime(d1, "%Y/%m/%d")
            end_date = start_date
        except ValueError:
            return {}
        t1, t2 = "08:30", "18:00"
    segments = {}
    current = start_date
    while current <= end_date:
        if start_date == end_date:
            start_t, end_t = t1, (t2 or "17:30")
        elif current == start_date:
            start_t, end_t = t1, "18:00"
        else:
            start_t, end_t = "08:30", "18:00"
        segments[current.date()] = (start_t, end_t)
        current += timedelta(days=1)
    return segments
def format_job_content(content):
    parts = re.split(r'[;；]', content)
    return [part.strip() for part in parts if part.strip()]
def find_excel_files():
    from source_processor import scan_sources
    return [str(p) for p in scan_sources('.')]

PROF_ORDER = [
    ("变电", ["变电"]),
    ("输电", ["输电"]),
    ("配电", ["配电"]),
    ("配网工程", ["配网工程", "配网", "配网施工", "配网检修", "配网运维"]),
    ("信通", ["信通", "通信", "信号", "信息通信"]),
    ("营销", ["营销", "业扩", "客户服务", "计量", "增容"])
]
def get_prof_priority(spec):
    if not spec:
        return len(PROF_ORDER)
    s = str(spec)
    for idx, (_, keys) in enumerate(PROF_ORDER):
        for k in keys:
            if k in s:
                return idx
    return len(PROF_ORDER)
def get_standard_profession(spec):
    if pd.isna(spec) or not str(spec).strip():
        return "未分类"
    s = str(spec).strip()
    if "营销" in s:
        return "营销"
    for prof_name, keys in PROF_ORDER:
        if prof_name == "营销":
            continue
        if any(k in s for k in keys):
            return prof_name
    marketing_keys = ["业扩", "客户服务", "高压计量", "高压增容", "计量", "增容"]
    if any(k in s for k in marketing_keys):
        return "营销"
    return s
def clean_text_value(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("nan", "none") else text
def get_level_priority(person_text):
    if not person_text:
        return 3
    s = str(person_text)
    if "领导" in s:
        return 1
    management_keywords = ["主任", "专责", "管理", "组长", "主管"]
    if any(k in s for k in management_keywords):
        return 2
    return 3
def clean_attachment_text(text):
    if pd.isna(text) or str(text).strip().lower() in ('nan', 'none', ''):
        return ""
    s = str(text).strip()
    s = re.sub(r"^附件(?:\s*[-—]?\s*\d+)+\s*[-—]?\s*", "", s)
    s = re.sub(r"^附件\s*[-—]?\s*", "", s)
    return s.strip()
def find_column_by_keywords(columns, keywords, required=True):
    for col in columns:
        col_clean = str(col).strip().replace('\n', '').replace(' ', '').replace('\r', '')
        for kw in keywords:
            if kw in col_clean:
                return col
    if required:
        raise KeyError(f"未找到匹配关键词 {keywords} 的列。")
    return None
def extract_risk_level(risk_val):
    if not risk_val:
        return None
    text = str(risk_val)
    for level in ("五级", "四级", "三级"):
        if level in text:
            return level
    return None
def infer_base_year_from_time_column(df, col_time):
    years = []
    for v in df[col_time].dropna():
        m = re.search(r'(\d{4})/\d{1,2}/\d{1,2}', str(v))
        if m:
            years.append(int(m.group(1)))
    return min(years) if years else datetime.now().year
def extract_date_range_from_title(title, base_year):
    m = re.search(r'(\d{1,2})月(\d{1,2})日(?:[—\-~～]|至)(\d{1,2})月(\d{1,2})日', str(title))
    if not m:
        raise ValueError("标题中未找到日期范围")
    m1, d1, m2, d2 = map(int, m.groups())
    start = datetime(base_year, m1, d1)
    end = datetime(base_year, m2, d2)
    if end < start:
        end = datetime(base_year + 1, m2, d2)
    return f"{m1}.{d1}", f"{m2}.{d2}", start, end
# ======================================
# 公共数据层
# ======================================
def load_source_df(src):
    wb = load_workbook(src, data_only=True)
    ws = wb.active
    title = str(ws["A1"].value).strip() if ws["A1"].value is not None else ""
    unmerge_and_fill(ws)
    headers = [cell.value for cell in ws[2]]
    data = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=3)]
    df = pd.DataFrame(data, columns=headers)
    df = df.dropna(how='all').reset_index(drop=True)
    wb.close()
    return df, title
def resolve_columns(df):
    cols = df.columns.tolist()
    return {
        'content': find_column_by_keywords(cols, ['作业内容', '工作内容', '作业项目', '作业任务']),
        'spec': find_column_by_keywords(cols, ['专业']),
        'risk': find_column_by_keywords(cols, ['风险等级', '作业风险', '风险级别']),
        'power_plan': find_column_by_keywords(cols, ['关联停电计划'], required=False),
        'time': df.columns[9] if len(df.columns) > 9 else None,
        'person': df.columns[17] if len(df.columns) > 17 else None,
        'project': df.columns[1] if len(df.columns) > 1 else None,
        'measures': find_column_by_keywords(
            cols, ['管控措施', '关键风险点及管控措施', '控制措施', '防控措施', '风险措施'], required=False
        ),
    }
def _safe_row_value(row, col):
    if col is None or col not in row.index:
        return None
    return row[col]
def build_standard_records(df, cols):
    col_content = cols['content']
    col_time = cols['time']
    if col_time is None:
        return []
    df_valid = df.dropna(subset=[col_content, col_time], how='any').reset_index(drop=True)
    records = []
    seen = set()
    for _, row in df_valid.iterrows():
        content_raw = clean_text_value(_safe_row_value(row, col_content))
        content_norm = normalize_content(_safe_row_value(row, col_content))
        time_raw = _safe_row_value(row, col_time)
        time_canon = canonicalize_time(time_raw)
        project_raw = clean_text_value(_safe_row_value(row, cols['project']))
        project_norm = normalize_content(_safe_row_value(row, cols['project']))
        project_display = project_raw[:-1] if project_raw.endswith("工程。") else project_raw
        spec_raw = clean_text_value(_safe_row_value(row, cols['spec']))
        spec_norm = normalize_content(_safe_row_value(row, cols['spec']))
        std_spec = get_standard_profession(spec_raw)
        is_marketing = std_spec == "营销"
        risk_raw = clean_text_value(_safe_row_value(row, cols['risk']))
        risk_norm = normalize_content(_safe_row_value(row, cols['risk']))
        risk_level = extract_risk_level(risk_raw)
        person_raw = clean_attachment_text(_safe_row_value(row, cols['person']))
        person_norm = normalize_content(_safe_row_value(row, cols['person']))
        measures_raw = clean_text_value(_safe_row_value(row, cols.get('measures')))
        measures_norm = normalize_content(_safe_row_value(row, cols.get('measures')))
        power_raw = clean_text_value(_safe_row_value(row, cols['power_plan']))
        power_norm = normalize_content(_safe_row_value(row, cols['power_plan']))
        has_power_plan = power_raw.strip().lower() not in ('否', '空', 'none', 'nan', '')
        if is_marketing:
            full_content = content_raw
        elif (project_display and "工程" in project_display
              and "：" not in project_display and ":" not in project_display
              and not content_raw.startswith(project_display + "：")
              and not content_raw.startswith(project_display + ":")):
            full_content = f"{project_display}：{content_raw}"
        else:
            full_content = content_raw
        l_rank = get_level_priority(person_raw)
        p_rank = get_prof_priority(spec_raw)
        is_low_voltage = "低压计量" in content_raw or "低压计量" in project_display
        dedup_key = (content_norm, time_canon, project_norm, spec_norm,
                     risk_norm, person_norm, power_norm)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        records.append({
            'content_raw': content_raw,
            'content_norm': content_norm,
            'full_content': full_content,
            'project_raw': project_raw,
            'project_display': project_display,
            'spec_raw': spec_raw,
            'std_spec': std_spec,
            'is_marketing': is_marketing,
            'risk_raw': risk_raw,
            'risk_level': risk_level,
            'time_raw': time_raw,
            'time_canon': time_canon,
            'person_raw': person_raw,
            'measures_raw': measures_raw,
            'has_power_plan': has_power_plan,
            'l_rank': l_rank,
            'p_rank': p_rank,
            'is_low_voltage': is_low_voltage,
        })
    return records
def expand_to_daily(records):
    daily = []
    for rec in records:
        segments = extract_daily_time_segments(rec['time_raw'])
        for seg_day, (st, et) in segments.items():
            daily.append({
                **rec,
                'date': seg_day,
                'start_time': st,
                'end_time': et,
            })
    return daily
# ======================================
# 源表预处理模块
# ======================================
def _has_strikethrough(cell):
    try:
        if cell.font and cell.font.strikethrough:
            return True
    except Exception:
        pass
    return False
def normalize_cell_line_breaks(ws, max_row, max_col):
    """统一换行符并清理空行（严格从第3行开始）"""
    changed = 0
    DATA_START_ROW = 3
    for row in range(DATA_START_ROW, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if cell.value is None or not isinstance(cell.value, str):
                continue
            original = cell.value
            if '\n' not in original and '\r' not in original:
                continue
            text = original.replace('\r\n', '\n').replace('\r', '\n')
            lines = [line.strip() for line in text.split('\n')]
            cleaned = []
            prev_empty = False
            for line in lines:
                if line == '':
                    if not prev_empty:
                        cleaned.append(line)
                    prev_empty = True
                else:
                    cleaned.append(line)
                    prev_empty = False
            while cleaned and cleaned[0] == '':
                cleaned.pop(0)
            while cleaned and cleaned[-1] == '':
                cleaned.pop()
            new_text = '\n'.join(cleaned)
            if new_text != original:
                cell.value = new_text
                changed += 1
    if changed > 0:
        safe_print(f"  📝 换行统一：修正了 {changed} 个单元格的换行格式")
    return changed
def merge_name_phone(ws, max_row, max_col):
    """合并单元格内换行的姓名与电话（无空格紧贴拼接，限定人员列）"""
    changed = 0
    DATA_START_ROW = 3
    phone_pattern = re.compile(r'^(1\d{10}|0\d{2,3}-?\d{7,8})$')
    target_cols = []
    for c in range(1, max_col + 1):
        header_val = str(ws.cell(2, c).value or "").strip()
        if any(kw in header_val for kw in ("人", "联系", "电话", "到岗", "监护")):
            target_cols.append(c)
    if not target_cols:
        target_cols = list(range(4, max_col + 1))
    for col in target_cols:
        for row in range(DATA_START_ROW, max_row + 1):
            cell = ws.cell(row, col)
            if not cell.value or not isinstance(cell.value, str):
                continue
            original = cell.value
            if '\n' not in original and '\r' not in original:
                continue
            text = original.replace('\r\n', '\n').replace('\r', '\n')
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            if len(lines) < 2:
                continue
            if any(phone_pattern.match(line) for line in lines):
                new_text = ''.join(lines)
                if new_text != original:
                    cell.value = new_text
                    changed += 1
    if changed > 0:
        safe_print(f"  📞 姓名电话合并：修正了 {changed} 个单元格（无空格紧贴拼接）")
    return changed
def fix_missing_time_dash(ws, max_row):
    """时间列破折号补齐：严格限定时间列"""
    DATA_START_ROW = 3
    max_col = min(ws.max_column, 20)
    time_col = None
    for r in (2, 1):
        for c in range(1, max_col + 1):
            val = str(ws.cell(r, c).value or "")
            if len(val) <= 20 and "时间" in val:
                time_col = c
                break
        if time_col:
            break
    if not time_col:
        return 0
    changed = 0
    for row in range(DATA_START_ROW, max_row + 1):
        cell = ws.cell(row, time_col)
        original = str(cell.value or "").strip()
        time_matches = list(re.finditer(r'\d{1,2}:\d{2}', original))
        if len(time_matches) < 2:
            continue
        first_end = time_matches[0].end()
        second_start = time_matches[1].start()
        between = original[first_end:second_start]
        has_separator = any(sep in between for sep in ('-', '–', '—', '至', '到', '~', '～'))
        if not has_separator:
            rest = original[first_end:]
            cell.value = original[:first_end] + '-' + rest.lstrip()
            changed += 1
    if changed > 0:
        safe_print(f"  ⏰ 时间补齐：修正了 {changed} 个时间单元格遗漏的\"-\"分隔符")
    return changed
_working_contexts = {}

def preprocess_source_file(src):
    from source_processor import working_copy, preprocess_sheet, Report, trim_inflated_copy
    context = working_copy(src)
    path = context.__enter__()
    _working_contexts[str(path)] = context
    wb = None
    try:
        report = Report(src)
        trim_inflated_copy(path, report)
        wb = load_workbook(path)
        preprocess_sheet(wb.worksheets[0], report)
        wb.save(path)
        for issue in report.issues:
            if issue['level'] in ('warn', 'error'):
                log_error(src, issue['code'], issue['cell'] + ' ' + issue['message'])
        if any(issue['level'] == 'error' for issue in report.issues):
            raise ValueError('源表存在日期、时间或必需列错误，请先修正报告中的问题再生成明细')
        return str(path)
    except Exception:
        cleanup_preprocessed_file(str(path))
        raise
    finally:
        if wb:
            wb.close()

def cleanup_preprocessed_file(path):
    context = _working_contexts.pop(str(path), None)
    if context:
        context.__exit__(None, None, None)

def _get_row_height(ws, row, default=15.0):
    h = ws.row_dimensions[row].height
    if h is None or h <= 0:
        return default
    return float(h)
def _estimate_text_height(text, col_width_chars, font_size=11):
    if not text:
        return 15.0
    lines = str(text).split('\n')
    total_lines = 0
    for line in lines:
        line_width = sum(2.3 if '\u4e00' <= ch <= '\u9fff' else 1.1 for ch in line)
        line_lines = max(1, int(math.ceil(line_width / col_width_chars)))
        total_lines += line_lines
    return max(15.0, total_lines * (font_size * 1.5 + 3) * 1.15)
def _cell_has_measures(cell):
    v = cell.value
    return bool(v) and ("管控措施" in str(v))
def _estimate_needed_height(cell, col_width_chars):
    if not cell.value or not str(cell.value).strip():
        return 0.0
    h = _estimate_text_height(cell.value, col_width_chars)
    if _cell_has_measures(cell):
        h += MEASURES_HEIGHT_PLUS
    return h
def insert_overflow_rows_by_measures(ws, start_row=3):
    from source_processor import column_map, format_sheet, Report
    header, cols = column_map(ws)
    return format_sheet(ws, header, cols, Report(), MEASURES_HEIGHT_PLUS)

def export_processed_source(cleaned_src_path, original_src):
    from source_processor import process_file
    # The processor creates its own protected work copy, including when called directly.
    target, _ = process_file(original_src, MEASURES_HEIGHT_PLUS)
    return str(target)

# ======================================
# 周计划明细插行与分页排版
# ======================================
def _get_merge_bounds(ws, row, col):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            is_anchor = (row == rng.min_row and col == rng.min_col)
            return is_anchor, rng.min_row, rng.min_col, rng.max_row, rng.max_col
    return True, row, col, row, col
def _apply_height_to_rows(ws, start_row, end_row, required_height):
    row_count = end_row - start_row + 1
    if row_count <= 0:
        return
    current_total = sum(_get_row_height(ws, r, default=15) for r in range(start_row, end_row + 1))
    if required_height <= current_total:
        return
    diff = required_height - current_total
    leftover = 0.0
    for offset in range(row_count):
        r = start_row + offset
        current = _get_row_height(ws, r, default=15)
        add_height = diff / row_count + leftover / (row_count - offset)
        if current + add_height > EXCEL_MAX_ROW_HEIGHT:
            actual_add = EXCEL_MAX_ROW_HEIGHT - current
            leftover += add_height - actual_add
        else:
            actual_add = add_height
            leftover = 0.0
        if actual_add > 0:
            ws.row_dimensions[r].height = current + actual_add
def insert_overflow_continuation_rows(ws, start_row, end_row, content_col_start, content_col_end, col_widths):
    total_inserted = 0
    continuation_rows = set()
    inserted_area_rows = set()
    max_col = content_col_end
    row = start_row
    while row <= end_row + total_inserted:
        max_needed_height = 0.0
        row_has_measures = False
        for col in range(content_col_start, content_col_end + 1):
            cell = ws.cell(row, col)
            if cell.value and str(cell.value).strip():
                w = col_widths.get(col, 25)
                h = _estimate_needed_height(cell, w)
                max_needed_height = max(max_needed_height, h)
                if _cell_has_measures(cell):
                    row_has_measures = True
        if max_needed_height > EXCEL_MAX_ROW_HEIGHT:
            rows_needed = int((max_needed_height + EXCEL_MAX_ROW_HEIGHT - 1) // EXCEL_MAX_ROW_HEIGHT)
            rows_to_insert = rows_needed - 1
            if rows_to_insert > 0:
                insert_at = row + 1
                end_insert_row = insert_at + rows_to_insert - 1
                ws.insert_rows(insert_at, amount=rows_to_insert)
                expanded_cols = set()
                for c in range(1, max_col + 1):
                    is_anchor, min_r, min_c, max_r, max_c = _get_merge_bounds(ws, row, c)
                    if not is_anchor or max_r < row:
                        continue
                    old_range = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
                    try:
                        ws.unmerge_cells(old_range)
                    except Exception:
                        pass
                    new_end_row = max(max_r, end_insert_row)
                    try:
                        ws.merge_cells(start_row=min_r, start_column=min_c,
                                       end_row=new_end_row, end_column=max_c)
                        expanded_cols.add(c)
                    except Exception:
                        pass
                for col in range(content_col_start, content_col_end + 1):
                    if col in expanded_cols:
                        continue
                    is_anchor, min_r, min_c, max_r, max_c = _get_merge_bounds(ws, row, col)
                    if min_r == row and max_r == row:
                        try:
                            ws.merge_cells(start_row=row, start_column=col,
                                           end_row=end_insert_row, end_column=col)
                        except Exception:
                            pass
                _apply_height_to_rows(ws, row, end_insert_row, max_needed_height)
                for c in range(1, max_col + 1):
                    is_anchor, min_r, min_c, max_r, max_c = _get_merge_bounds(ws, row, c)
                    if min_r <= row <= max_r and min_c <= c <= max_c and (max_r > row or max_c > c):
                        continue
                    src_cell = ws.cell(row, c)
                    for r in range(row + 1, end_insert_row + 1):
                        dst_cell = ws.cell(r, c)
                        if src_cell.has_style:
                            dst_cell.font = src_cell.font.copy()
                            dst_cell.border = src_cell.border.copy()
                            dst_cell.fill = src_cell.fill.copy()
                            dst_cell.alignment = src_cell.alignment.copy()
                safe_print(f"  📏 明细插行续行：第{row}行内容超长(约{max_needed_height:.0f}磅)，插入{rows_to_insert}行续行")
                total_inserted += rows_to_insert
                for r in range(row + 1, end_insert_row + 1):
                    continuation_rows.add(r)
                for r in range(row, end_insert_row + 1):
                    inserted_area_rows.add(r)
                row = end_insert_row + 1
                continue
        else:
            if max_needed_height > 15.0:
                ws.row_dimensions[row].height = max_needed_height
            elif row_has_measures and MEASURES_HEIGHT_PLUS:
                ws.row_dimensions[row].height = 15.0 + MEASURES_HEIGHT_PLUS
        row += 1
    return total_inserted, continuation_rows, inserted_area_rows
def apply_no_split_page_breaks(ws, start_row, end_row):
    page_height = A3_LANDSCAPE_PRINTABLE_HEIGHT * PAGE_HEIGHT_SAFETY_FACTOR
    title_height = _get_row_height(ws, 1, default=45)
    groups = []
    row = start_row
    while row <= end_row:
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        if a_val and not b_val:
            h = _get_row_height(ws, row, default=15)
            groups.append((row, row, h))
            row += 1
        else:
            group_end = min(row + 2, end_row)
            h = sum(_get_row_height(ws, r, default=25) for r in range(row, group_end + 1))
            groups.append((row, group_end, h))
            row = group_end + 1
    current_height = title_height
    break_rows = []
    for g_start, g_end, g_height in groups:
        if current_height + g_height > page_height and current_height > title_height + 10:
            break_rows.append(g_start)
            current_height = title_height + g_height
        else:
            current_height += g_height
    ws.row_breaks = RowBreak()
    for br in break_rows:
        ws.row_breaks.append(Break(id=br))
    if break_rows:
        safe_print(f"  📄 不跨行分页：已在 {len(break_rows)} 处插入分页符")
    return break_rows
def setup_a3_print(ws, total_rows, total_cols):
    ws.page_setup.paperSize = 8
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.gridLines = False
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.15
    ws.page_margins.bottom = 0.15
    ws.page_margins.header = 0.05
    ws.page_margins.footer = 0.05
    last_col_letter = get_column_letter(total_cols)
    ws.print_area = f"A1:{last_col_letter}{total_rows}"
    ws.print_title_rows = '1:1'
# ======================================
# 统计与主任务生成
# ======================================
def print_statistics(daily_data, start_date, end_date):
    print("\n📊 统计概况:")
    total_items = sum(len(plans) for plans in daily_data.values())
    print(f"📅 日期范围: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}")
    print(f"🔢 总计划项数: {total_items}")
    print("\n📆 每日计划分布:")
    total_daily_plans = 0
    for i in range((end_date - start_date).days + 1):
        day = start_date + timedelta(days=i)
        count = len(daily_data.get(day, []))
        total_daily_plans += count
        weekday = "一二三四五六日"[day.weekday()]
        print(f"  {day.strftime('%m-%d')} 星期{weekday}: {count}项计划")
    print(f"\n📈 本周计划总数（按天统计）: {total_daily_plans}项")
def generate_weekly_plan_v5(all_files):
    safe_print("\n📝 正在生成周计划明细（V5增强版）...")
    if not all_files:
        safe_print("⚠️ 没有找到可处理的文件")
        return
    for src in all_files:
        processed_src = None
        try:
            processed_src = preprocess_source_file(src)
            safe_print(f"\n📁 读取文件: {src}")
            df, title = load_source_df(processed_src)
            if df.empty:
                safe_print("⚠️ 未找到有效数据")
                continue
            cols = resolve_columns(df)
            if cols['time'] is None:
                safe_print("⚠️ 时间列获取失败")
                continue
            records = build_standard_records(df, cols)
            daily_records = expand_to_daily(records)
            base_year = infer_base_year_from_time_column(df, cols['time'])
            try:
                start_short, end_short, start_date, end_date = extract_date_range_from_title(title, base_year)
            except Exception as e:
                safe_print(f"⚠️ 解析日期范围失败: {e}")
                continue
            daily_data = {}
            for rec in daily_records:
                if rec['is_low_voltage']:
                    continue
                day_dt = datetime.combine(rec['date'], datetime.min.time())
                if not (start_date <= day_dt <= end_date):
                    continue
                job_desc_parts = [
                    rec['full_content'],
                    f"专业：{rec['spec_raw']}",
                    f"作业风险等级：{rec['risk_raw']}",
                    f"作业时间：{rec['start_time']}-{rec['end_time']}"
                ]
                if rec['has_power_plan']:
                    job_desc_parts.append("【关联月度停电计划】")
                job_desc = "\n".join(job_desc_parts)
                daily_data.setdefault(day_dt, []).append({
                    "person": rec['person_raw'],
                    "job_desc": job_desc,
                    "spec": rec['spec_raw'],
                    "risk_val": rec['risk_raw'],
                    "has_measures": False,
                    "l_rank": rec['l_rank'],
                    "p_rank": rec['p_rank'],
                })
            print_statistics(daily_data, start_date, end_date)
            generate_excel_output_v5(daily_data, start_date, end_date, src)
            safe_print(f"✅ 周计划明细完成: {src}")
        except Exception as e:
            log_error(src, "周计划明细生成失败(V5)", str(e), traceback.format_exc())
            safe_print(f"❌ 文件处理失败: {src} | {e}", error=True)
        finally:
            cleanup_preprocessed_file(processed_src)

def generate_excel_output_v5(daily_data, start_date, end_date, src):
    PLANS_PER_GROUP = 7
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "周计划明细"
    total_cols = PLANS_PER_GROUP + 1
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    no_border = Border()
    header_font = Font(bold=True, name="微软雅黑")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="微软雅黑")
    center_align = Alignment(wrap_text=True, horizontal="center", vertical="center")
    left_align = Alignment(wrap_text=True, horizontal="left", vertical="center")
    title_text = f"{start_date.month}月{start_date.day}日-{end_date.month}月{end_date.day}日计划明细表"
    ws_out.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws_out.cell(1, 1, title_text)
    title_cell.font = Font(size=18, bold=True, name="微软雅黑")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_out.row_dimensions[1].height = 45
    for col_idx in range(1, total_cols + 1):
        ws_out.cell(1, col_idx).border = no_border
    current_row = 2
    data_start_row = 2
    num_days = (end_date - start_date).days + 1
    for i in range(num_days):
        day = start_date + timedelta(days=i)
        weekday_idx = (start_date.weekday() + i) % 7
        weekday = "一二三四五六日"[weekday_idx]
        label = f"星期{weekday} {day.month}月{day.day}日"
        day_plans = daily_data.get(day, [])
        day_plans.sort(key=lambda x: (x["l_rank"], x["p_rank"]))
        if not day_plans:
            date_cell = ws_out.cell(current_row, 1, label)
            date_cell.font = data_font
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, total_cols + 1):
                ws_out.cell(current_row, col_idx).border = border
            current_row += 1
            continue
        num_groups = (len(day_plans) + PLANS_PER_GROUP - 1) // PLANS_PER_GROUP
        total_day_rows = num_groups * 3
        date_start_row = current_row
        date_end_row = current_row + total_day_rows - 1
        ws_out.merge_cells(start_row=date_start_row, start_column=1,
                          end_row=date_end_row, end_column=1)
        date_cell = ws_out.cell(date_start_row, 1, label)
        date_cell.font = data_font
        date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in range(date_start_row, date_end_row + 1):
            top_side = thin if r == date_start_row else None
            bottom_side = thin if r == date_end_row else None
            ws_out.cell(r, 1).border = Border(left=thin, right=thin, top=top_side, bottom=bottom_side)
        for group_idx in range(num_groups):
            group_start = group_idx * PLANS_PER_GROUP
            group_end = min(group_start + PLANS_PER_GROUP, len(day_plans))
            group_size = group_end - group_start
            group_plans = day_plans[group_start:group_end]
            header_row = current_row
            ws_out.row_dimensions[header_row].height = 25
            for j in range(group_size):
                col_idx = j + 2
                plan_label = f"计划{group_start + j + 1}"
                header_cell = ws_out.cell(header_row, col_idx, plan_label)
                header_cell.font = header_font
                header_cell.alignment = header_alignment
                header_cell.border = border
            for col_idx in range(group_size + 2, total_cols + 1):
                ws_out.cell(header_row, col_idx).border = border
            current_row += 1
            person_row = current_row
            for j in range(group_size):
                col_idx = j + 2
                plan = group_plans[j]
                person_cell = ws_out.cell(person_row, col_idx)
                person_cell.value = plan["person"]
                person_cell.font = data_font
                person_cell.alignment = center_align
                person_cell.border = border
            for col_idx in range(group_size + 2, total_cols + 1):
                ws_out.cell(person_row, col_idx).border = border
            current_row += 1
            job_row = current_row
            for j in range(group_size):
                col_idx = j + 2
                plan = group_plans[j]
                job_cell = ws_out.cell(job_row, col_idx)
                job_cell.value = plan["job_desc"]
                job_cell.font = data_font
                job_cell.alignment = left_align
                job_cell.border = border
            for col_idx in range(group_size + 2, total_cols + 1):
                ws_out.cell(job_row, col_idx).border = border
            current_row += 1
    data_end_row = current_row - 1
    ws_out.column_dimensions["A"].width = 18
    col_widths = {}
    for col in range(2, total_cols + 1):
        col_letter = get_column_letter(col)
        ws_out.column_dimensions[col_letter].width = 25
        col_widths[col] = 25
    safe_print("  📏 正在为周计划明细插行续行并设置行高...")
    inserted, _, _ = insert_overflow_continuation_rows(
        ws_out, data_start_row, data_end_row,
        content_col_start=2, content_col_end=total_cols,
        col_widths=col_widths
    )
    data_end_row += inserted
    safe_print("  📄 正在计算明细表不跨行分页位置...")
    apply_no_split_page_breaks(ws_out, data_start_row, data_end_row)
    setup_a3_print(ws_out, data_end_row, total_cols)
    from source_processor import output_folder, detail_name
    folder = output_folder(os.path.dirname(os.path.abspath(src)))
    folder.mkdir(parents=True, exist_ok=True)
    out_name = folder / detail_name(src, str(ws_out['A1'].value or ''))
    wb_out.save(out_name)
    wb_out.close()
    safe_print(f"💾 已保存周计划明细: {out_name}")
def generate_risk_txt(all_files):
    safe_print("\n📝 正在生成三四五级风险汇总 TXT...")
    if not all_files:
        safe_print("⚠️ 没有找到可处理的文件")
        return
    risk_data = {"三级": [], "四级": [], "五级": []}
    for src in all_files:
        processed_src = None
        try:
            processed_src = preprocess_source_file(src)
            df, _ = load_source_df(processed_src)
            if df.empty:
                continue
            cols = resolve_columns(df)
            records = build_standard_records(df, cols)
            daily_records = expand_to_daily(records)
            for rec in daily_records:
                risk_level = rec['risk_level']
                if not risk_level:
                    continue
                risk_data[risk_level].append((rec['date'], rec['full_content']))
        except Exception as e:
            log_error(src, "三四五级TXT生成失败", str(e), traceback.format_exc())
        finally:
            cleanup_preprocessed_file(processed_src)
    lines = []
    for level in ["三级", "四级", "五级"]:
        lines.append(f"{level}作业风险管控情况\n")
        jobs = sorted(risk_data[level], key=lambda x: x[0])
        if not jobs:
            lines.append(f"暂无{level}风险作业\n\n")
            continue
        date_groups = OrderedDict()
        for d, job in jobs:
            date_groups.setdefault(d, []).append(job)
        for day_dt, jobs in date_groups.items():
            lines.append(f"{day_dt.month}月{day_dt.day}日")
            for job in jobs:
                for part in format_job_content(job):
                    lines.append(part)
            lines.append("")
    txt_path = "三四五级风险汇总(简化版).txt"
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))
    safe_print(f"💾 已保存: {os.path.abspath(txt_path)}")
def generate_summary_txt(all_files):
    safe_print("\n📝 正在生成周计划项目汇总...")
    if not all_files:
        safe_print("⚠️ 没有找到可处理的文件")
        return
    total_count = 0
    spec_counts = {}
    risk_counts = {}
    yx_high_count = 0
    yx_low_count = 0
    for src in all_files:
        processed_src = None
        try:
            processed_src = preprocess_source_file(src)
            df, _ = load_source_df(processed_src)
            if df.empty:
                continue
            cols = resolve_columns(df)
            records = build_standard_records(df, cols)
            for rec in records:
                total_count += 1
                std_spec = rec['std_spec']
                spec_counts[std_spec] = spec_counts.get(std_spec, 0) + 1
                risk = rec['risk_level'] or rec['risk_raw']
                if risk:
                    risk_counts[risk] = risk_counts.get(risk, 0) + 1
                if std_spec == '营销':
                    content = rec['content_raw']
                    raw_spec = rec['spec_raw']
                    high_keywords = ['高压计量', '高压业扩', '业扩', '高压增容', '增容']
                    if any(kw in raw_spec or kw in content for kw in high_keywords):
                        yx_high_count += 1
                    elif '低压计量' in raw_spec or '低压计量' in content:
                        yx_low_count += 1
        except Exception as e:
            log_error(src, "汇总统计生成失败", str(e), traceback.format_exc())
        finally:
            cleanup_preprocessed_file(processed_src)
    std_prof_names = [name for name, _ in PROF_ORDER]
    spec_parts = []
    for prof_name in std_prof_names:
        if prof_name not in spec_counts:
            continue
        if prof_name == '营销':
            detail_parts = []
            if yx_high_count:
                detail_parts.append(f"高压类{yx_high_count}项")
            if yx_low_count:
                detail_parts.append(f"低压计量{yx_low_count}项")
            detail = f"（{'、'.join(detail_parts)}）" if detail_parts else ""
            spec_parts.append(f"营销专业{spec_counts[prof_name]}项{detail}")
        else:
            spec_parts.append(f"{prof_name}专业{spec_counts[prof_name]}项")
    for prof_name, cnt in spec_counts.items():
        if prof_name not in std_prof_names:
            spec_parts.append(f"{prof_name}专业{cnt}项")
    spec_text = "、".join(spec_parts) if spec_parts else "暂无专业分类"
    risk_parts = []
    for risk in ('三级', '四级', '五级'):
        if risk in risk_counts:
            risk_parts.append(f"{risk}作业风险{risk_counts[risk]}项")
    for risk, cnt in risk_counts.items():
        if risk not in ('三级', '四级', '五级'):
            risk_parts.append(f"{risk}{cnt}项")
    risk_text = "、".join(risk_parts) if risk_parts else "无三级、四级、五级风险作业"
    summary_text = f"本周共安排作业计划{total_count}项，包括{spec_text}，其中{risk_text}。"
    txt_path = "周计划项目汇总.txt"
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(summary_text)
    safe_print(f"\n📊 汇总结果: {summary_text}")
    safe_print(f"💾 已保存: {os.path.abspath(txt_path)}")
def generate_processed_source_files(all_files):
    """独立执行：对源表管控措施超长进行全列贯穿插行，生成《处理后的源表》"""
    safe_print("\n📝 正在生成【处理后的源表】（管控措施超长整行插行）...")
    if not all_files:
        safe_print("⚠️ 没有找到可处理的源文件")
        return
    for src in all_files:
        try:
            path = export_processed_source(None, src)
            safe_print(f"已生成：{path}")
        except Exception as exc:
            log_error(src, "OUTPUT_FAILED", str(exc))
            safe_print(f"处理失败：{src} | {exc}", error=True)

def cleanup_error_log():
    if _error_count == 0 and os.path.exists(ERROR_LOG_PATH):
        try:
            os.remove(ERROR_LOG_PATH)
        except Exception:
            pass
# -------------------- 主程序菜单 --------------------
def main():
    safe_print("📌 周计划工具（V5功能移植版 - 稳定修复版）")
    safe_print("====================")
    safe_print("功能选项:")
    safe_print("1. 周计划明细 Excel 输出（V5增强版）")
    safe_print("2. 三四五级风险汇总 TXT 输出")
    safe_print("3. 周计划项目汇总文本输出")
    safe_print("4. 导出处理后的源表（管控措施超长全列插行）")
    safe_print("5. 退出")
    choice = input("请输入数字选择功能（可多选，逗号分隔，如1,2,3,4）: ")
    choices = [c.strip() for c in re.split(r"[,，、\s]+", choice) if c.strip()]
    all_files = find_excel_files()
    if "1" in choices:
        generate_weekly_plan_v5(all_files)
    if "2" in choices:
        generate_risk_txt(all_files)
    if "3" in choices:
        generate_summary_txt(all_files)
    if "4" in choices:
        generate_processed_source_files(all_files)
    cleanup_error_log()
    safe_print("\n✅ 所有任务完成")
    try:
        input("\n按回车键退出...")
    except Exception:
        pass
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        safe_print(f"\n❌ 程序异常退出: {e}", error=True)
        import traceback
        traceback.print_exc()
        try:
            input("\n按回车键退出...")
        except Exception:
            pass
